import logging
import os
import json
import csv
import re
import zipfile
import io
import uuid
from io import StringIO, BytesIO
from datetime import datetime, timedelta, time
from datetime import date
from dotenv import load_dotenv
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from telegram import (
    Update,
    ReplyKeyboardMarkup,
    KeyboardButton,
    InlineKeyboardButton,
    InlineKeyboardMarkup
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters
)
from telegram.request import HTTPXRequest

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)
# ========== Константы ==========
# Константы файлов
INGREDIENTS_FILE = "ingredients.json"
RECIPES_FILE = "recipes.json"
SETTINGS_FILE = "settings.json"
SALES_FILE = "sales.json"
PLANS_FILE = "plans.json"
CUSTOMERS_FILE = "customers.json"
ORDERS_FILE = "orders.json"
WRITEOFFS_FILE = "writeoffs.json"
PRICE_HISTORY_FILE = "price_history.json"
BATCHES_FILE = "batches.json"
# Состояния для онбординга
ONBOARDING_START, ONBOARDING_INGREDIENT, ONBOARDING_RECIPE, ONBOARDING_PRICE, ONBOARDING_FINISH = range(5)
# ========== Тексты для разделов помощи ==========
HELP_MAIN = (
    "📋 Основные команды\n\n"
    "/start - показать главное меню\n"
    "/menu - показать меню\n"
    "/help - показать это сообщение"
)

HELP_INGREDIENTS = (
    "📦 Ингредиенты и закупки\n\n"
    "/add_ingredient название цена единица - добавить ингредиент (или цена_упаковки вес единица)\n"
    "  Пример: /add_ingredient мука 50 кг\n"
    "  Пример с упаковкой: /add_ingredient масло 209.99 180 г\n"
    "/ingredients - список ингредиентов\n"
    "/remove_ingredient название - удалить ингредиент\n"
    "  Пример: /remove_ingredient мука\n"
    "/update_price название новая_цена - изменить цену ингредиента\n"
    "  Пример: /update_price мука 55\n"
    "/add_stock название количество - добавить остатки на склад\n"
    "  Пример: /add_stock мука 10\n"
    "/stock - показать текущие остатки\n"
    "/low_stock [порог] - показать ингредиенты с остатком ниже порога\n"
    "  Пример: /low_stock 2\n"
    "/purchase ингредиент количество цена ГГГГ-ММ-ДД [поставщик] - зарегистрировать закупку с указанием срока годности\n"
    "  Пример: /purchase мука 10 500 2026-06-01 Мельница\n"
    "/set_shelf_life ингредиент общий_срок [срок_после_вскрытия] - установить срок годности для ингредиента\n"
    "  Пример: /set_shelf_life сливки 14 5\n"
    "/open ингредиент [batch_id] - отметить партию как открытую\n"
    "  Пример: /open сливки\n"
    "/expiring [дней] - показать продукты, у которых срок истекает в ближайшие N дней\n"
    "  Пример: /expiring 7"
)

HELP_RECIPES = (
    "🍰 Рецепты\n\n"
    "/add_recipe Название: порции; ингредиенты - добавить рецепт (старый формат)\n"
    "  Пример: /add_recipe Омлет: 2; яйца 3, молоко 0.1\n"
    "/add_recipe2 название тип базовое_количество: ингредиенты - добавить рецепт для масштабирования\n"
    "  Пример: /add_recipe2 торт вес 1: мука 0.5, сахар 0.2, яйца 3\n"
    "/recipes [категория] - список рецептов (можно указать категорию)\n"
    "  Пример: /recipes рулеты\n"
    "/remove_recipe название - удалить рецепт\n"
    "  Пример: /remove_recipe омлет\n"
    "/delete_recipes - удалить все рецепты\n"
    "/show_recipe название - показать рецепт с описанием и расчётами\n"
    "  Пример: /show_recipe меренговый_рулет_белый\n"
    "/set_description название описание - добавить описание к рецепту\n"
    "  Пример: /set_description меренговый_рулет_белый Взбить белки с сахаром...\n"
    "/scale название новое_количество [единица] - пересчитать рецепт\n"
    "  Пример: /scale торт 2.5 кг\n"
    "  Пример: /scale рулет 3 шт\n"
    "/calculate название - рассчитать себестоимость ингредиентов\n"
    "  Пример: /calculate меренговый_рулет_белый\n"
    "/price_list - список всех рецептов с ценами продажи"
)

HELP_SALES = (
    "💰 Продажи и аналитика\n\n"
    "/use название [количество] [факт_цена] - списать ингредиенты и зафиксировать продажу\n"
    "  Пример: /use рулет 2 (цена по наценке)\n"
    "  Пример: /use рулет 2 1500 (общая сумма продажи)\n"
    "/order название_рецепта количество [цена_продажи] - добавить заказ (оценка, без списания)\n"
    "  Пример: /order рулет 2 (цена по наценке)\n"
    "  Пример: /order рулет 2 1500 (с указанием цены)\n"
    "/stats [день|неделя|месяц|год] - статистика продаж за период\n"
    "  Пример: /stats месяц\n"
    "/popular - топ-5 рецептов по продажам"
)

HELP_WRITEOFFS = (
    "🔄 Возвраты и списания\n\n"
    "/write_off название количество [причина] - списать ингредиенты или готовую продукцию\n"
    "  Пример: /write_off мука 0.5 просыпалась\n"
    "  Пример: /write_off меренговый_рулет_белый 2 брак\n"
    "/refund клиент рецепт количество [дата] [причина] - оформить возврат от клиента\n"
    "  Пример: /refund Иван меренговый_рулет_белый 1 не понравился\n"
    "  Пример с датой: /refund Иван меренговый_рулет_белый 1 2026-03-15 не понравился"
)

HELP_CUSTOMERS = (
    "📅 Клиенты и заказы\n\n"
    "/add_customer имя телефон [адрес] - добавить клиента\n"
    "  Пример: /add_customer Иван +79991234567\n"
    "/preorder клиент рецепт количество ГГГГ-ММ-ДД - создать предзаказ\n"
    "  Пример: /order Иван меренговый_рулет_белый 2 2026-03-20\n"
    "/orders [ГГГГ-ММ-ДД] - показать заказы на дату\n"
    "  Пример: /orders 2026-03-20\n"
    "/remind - напомнить о заказах на сегодня и завтра"
)

HELP_PLANS = (
    "📊 Планирование закупок и категории\n\n"
    "/plan рецепт количество ГГГГ-ММ-ДД - запланировать приготовление\n"
    "  Пример: /plan меренговый_рулет_белый 5 2026-03-20\n"
    "/shopping_list [ГГГГ-ММ-ДД] - список необходимых закупок на дату (или по всем планам)\n"
    "  Пример: /shopping_list 2026-03-20\n"
    "/set_category рецепт категория - присвоить категорию рецепту\n"
    "  Пример: /set_category меренговый_рулет_белый рулеты\n"
    "/categories - список всех категорий\n"
    "/recipes категория - показать рецепты только указанной категории\n"
    "  Пример: /recipes рулеты"
)

HELP_IMPORT_EXPORT = (
    "📁 Импорт/экспорт данных\n\n"
    "/export - выгрузить ингредиенты в CSV\n"
    "/export_xlsx - выгрузить ингредиенты в Excel (формат XLSX)\n"
    "/export_full - выгрузить все данные в ZIP-архив (JSON-файлы)\n"
    "/export_full_excel - выгрузить все данные в один Excel-файл (несколько листов)\n"
    "/report_xlsx [день|неделя|месяц|год] - выгрузить отчёт по заказам в Excel\n"
    "  Пример: /report_xlsx месяц\n"
    "/import_recipe - импортировать рецепт из текста (пошаговый диалог)\n"
    "/parse - упрощённый парсер рецепта (без диалога)\n"
    "  Пример: /parse мука 200 г, сахар 150 г"
)

HELP_ADVANCED = (
    "⚙️ Дополнительные команды\n\n"
    "/set_hourly_rate ставка - установить почасовую ставку работы (глобально)\n"
    "  Пример: /set_hourly_rate 350\n"
    "/set_packaging название цена - установить стоимость упаковки для рецепта\n"
    "  Пример: /set_packaging меренговый_рулет_белый 76\n"
    "/set_work_hours название часы - установить время работы на рецепт\n"
    "  Пример: /set_work_hours меренговый_рулет_белый 0.75\n"
    "/set_markup название процент - установить наценку для рецепта\n"
    "  Пример: /set_markup меренговый_рулет_белый 40\n"
    "/price_history название - показать историю изменения цены ингредиента\n"
    "  Пример: /price_history мука"
)
def back_to_help_menu():
    """Создаёт клавиатуру с одной кнопкой для возврата в меню помощи"""
    keyboard = [[InlineKeyboardButton("« Назад к разделам", callback_data="help_back")]]
    return InlineKeyboardMarkup(keyboard)
def load_data(filename):
    if os.path.exists(filename):
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_data(data, filename):
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
# Глобальные переменные
ingredients = {}
recipes = {}
settings = {}
sales = []
plans = []
customers = {}
orders = []
batches = []
price_history = []  # глобальный список партий
writeoffs = []  # список списаний и возвратов
# Состояния для диалога импорта рецепта
WAITING_RECIPE_TEXT, WAITING_INGREDIENT_PRICE, WAITING_RECIPE_NAME, WAITING_RECIPE_TYPE = range(4)
# ========== Функции для работы с файлами ==========
def load_settings():
    global settings
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                settings = json.load(f)
        except:
            settings = {}
    else:
        settings = {}
    if 'hourly_rate' not in settings:
        settings['hourly_rate'] = 0.0
        save_settings()
def load_batches():
    global batches
    if os.path.exists(BATCHES_FILE):
        try:
            with open(BATCHES_FILE, 'r', encoding='utf-8') as f:
                batches = json.load(f)
        except:
            batches = []
    else:
        batches = []

def save_batches():
    with open(BATCHES_FILE, 'w', encoding='utf-8') as f:
        json.dump(batches, f, ensure_ascii=False, indent=2)
def save_settings():
    with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)

def load_sales():
    global sales
    if os.path.exists(SALES_FILE):
        try:
            with open(SALES_FILE, 'r', encoding='utf-8') as f:
                sales = json.load(f)
        except:
            sales = []
    else:
        sales = []

def save_sales():
    with open(SALES_FILE, 'w', encoding='utf-8') as f:
        json.dump(sales, f, ensure_ascii=False, indent=2)

def load_plans():
    global plans
    if os.path.exists(PLANS_FILE):
        try:
            with open(PLANS_FILE, 'r', encoding='utf-8') as f:
                plans = json.load(f)
        except:
            plans = []
    else:
        plans = []

def save_plans():
    with open(PLANS_FILE, 'w', encoding='utf-8') as f:
        json.dump(plans, f, ensure_ascii=False, indent=2)

def load_customers():
    global customers
    if os.path.exists(CUSTOMERS_FILE):
        try:
            with open(CUSTOMERS_FILE, 'r', encoding='utf-8') as f:
                customers = json.load(f)
        except:
            customers = {}
    else:
        customers = {}

def save_customers():
    with open(CUSTOMERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(customers, f, ensure_ascii=False, indent=2)

def load_orders():
    global orders
    if os.path.exists(ORDERS_FILE):
        try:
            with open(ORDERS_FILE, 'r', encoding='utf-8') as f:
                orders = json.load(f)
        except:
            orders = []
    else:
        orders = []

def save_orders():
    with open(ORDERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(orders, f, ensure_ascii=False, indent=2)
# ========== Команда /order (новый заказ) ==========
# ========== Команда /order (новый заказ) ==========
# ========== Команда /order (новый заказ) ==========
async def order_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Добавляет новый заказ: /order название_рецепта количество [цена_продажи]"""
    args = context.args
    if len(args) < 2:
        await update.message.reply_text(
            "Формат: /order название_рецепта количество [цена_продажи]\n"
            "Примеры:\n"
            "/order меренговый_рулет_фисташковый 2\n"
            "/order меренговый_рулет_фисташковый 2 3500"
        )
        return

    # Парсим аргументы
    price = None
    qty = None
    name_parts = args

    try:
        last_arg = args[-1].replace(',', '.')
        float(last_arg)
        if len(args) >= 3:
            try:
                qty = float(args[-2].replace(',', '.'))
                price = float(last_arg)
                name_parts = args[:-2]
            except ValueError:
                qty = float(last_arg)
                name_parts = args[:-1]
        else:
            qty = float(last_arg)
            name_parts = args[:-1]
    except ValueError:
        qty = 1.0
        name_parts = args

    name = ' '.join(name_parts).lower()

    if name not in recipes:
        await update.message.reply_text(f"Рецепт '{name}' не найден.")
        return

    data = recipes[name]

    # --- Расчёт себестоимости ---
    if isinstance(data, dict) and "type" in data:
        ing_dict = data["ingredients"]
        base_qty = data.get("base_qty", 1)
        scale = qty / base_qty if data["type"] == "weight" else qty
    else:
        ing_dict = data.get("ingredients", data) if isinstance(data, dict) else data
        portions = data.get("portions", 1) if isinstance(data, dict) else 1
        scale = qty / portions if portions != 1 else qty

    scaled = {ing: ing_qty * scale for ing, ing_qty in ing_dict.items()}

    total_ing = 0.0
    missing = []
    for ing_name, ing_qty in scaled.items():
        if ing_name in ingredients:
            total_ing += ingredients[ing_name]["price"] * ing_qty
        else:
            missing.append(ing_name)

    if missing:
        await update.message.reply_text(f"❌ Не хватает ингредиентов в базе: {', '.join(missing)}")
        return

    packaging = data.get('packaging', 0.0) if isinstance(data, dict) else 0.0
    work_hours = data.get('work_hours', 0.0) if isinstance(data, dict) else 0.0
    hourly_rate = settings.get('hourly_rate', 0.0)
    work_cost = work_hours * hourly_rate * (qty if 'type' in data and data['type'] == 'pcs' else 1)

    total_cost = total_ing + packaging * (qty if 'type' in data and data['type'] == 'pcs' else 1) + work_cost

    if price is None:
        markup = data.get('markup', 50) if isinstance(data, dict) else 50
        price = total_cost * (1 + markup / 100)
        price = round(price)

    profit = price - total_cost

    # Запись заказа
    order_record = {
        "date": datetime.now().strftime("%Y-%m-%d"),
        "recipe": name,
        "quantity": qty,
        "cost": total_cost,
        "revenue": price,
        "profit": profit
    }
    orders.append(order_record)
    save_orders()

    # Формируем ответ (без Markdown)
    msg = (
        f"📦 Заказ добавлен\n\n"
        f"Десерт: {name}\n"
        f"Количество: {qty} шт\n\n"
        f"Себестоимость: {total_cost:.0f} ₽\n"
        f"Цена продажи: {price:.0f} ₽\n\n"
        f"💰 Прибыль: {profit:.0f} ₽"
    )
    await update.message.reply_text(msg)
# ========== Команда /profit ==========
async def profit_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает прибыль за сегодня и месяц (без Markdown)"""
    today = datetime.now().date()
    start_of_month = datetime(today.year, today.month, 1).date()

    today_orders = []
    month_orders = []

    for order in orders:
        # Проверяем, есть ли ключ 'date' (новый формат) или 'due_date' (старый формат)
        if 'date' in order:
            order_date = datetime.strptime(order['date'], "%Y-%m-%d").date()
        elif 'due_date' in order:
            order_date = datetime.strptime(order['due_date'], "%Y-%m-%d").date()
        else:
            continue  # пропускаем, если нет даты

        if order_date == today:
            today_orders.append(order)
        if order_date >= start_of_month:
            month_orders.append(order)

    def calc_stats(order_list):
        count = len(order_list)
        revenue = sum(o.get('revenue', o.get('price', 0)) for o in order_list)
        cost = sum(o.get('cost', 0) for o in order_list)
        profit = revenue - cost
        return count, revenue, cost, profit

    today_count, today_rev, today_cost, today_profit = calc_stats(today_orders)
    month_count, month_rev, month_cost, month_profit = calc_stats(month_orders)

    msg = (
        f"📊 Прибыль\n\n"
        f"Сегодня\n"
        f"Заказы: {today_count}\n"
        f"Выручка: {today_rev:.0f} ₽\n"
        f"Себестоимость: {today_cost:.0f} ₽\n"
        f"Прибыль: {today_profit:.0f} ₽\n\n"
        f"Месяц\n"
        f"Заказы: {month_count}\n"
        f"Выручка: {month_rev:.0f} ₽\n"
        f"Себестоимость: {month_cost:.0f} ₽\n"
        f"Прибыль: {month_profit:.0f} ₽"
    )
    await update.message.reply_text(msg)
def load_writeoffs():
    global writeoffs
    if os.path.exists(WRITEOFFS_FILE):
        try:
            with open(WRITEOFFS_FILE, 'r', encoding='utf-8') as f:
                writeoffs = json.load(f)
        except:
            writeoffs = []
    else:
        writeoffs = []

def save_writeoffs():
    with open(WRITEOFFS_FILE, 'w', encoding='utf-8') as f:
        json.dump(writeoffs, f, ensure_ascii=False, indent=2)
def save_orders():
    with open(ORDERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(orders, f, ensure_ascii=False, indent=2)
def load_price_history():
    global price_history
    if os.path.exists(PRICE_HISTORY_FILE):
        try:
            with open(PRICE_HISTORY_FILE, 'r', encoding='utf-8') as f:
                price_history = json.load(f)
        except:
            price_history = []
    else:
        price_history = []

def save_price_history():
    with open(PRICE_HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(price_history, f, ensure_ascii=False, indent=2)

# ========== Меню с кнопками ==========


def get_main_keyboard():
    keyboard = [
        [KeyboardButton("➕ Добавить ингредиент")],
        [KeyboardButton("📋 Список ингредиентов")],
        [KeyboardButton("🍰 Добавить рецепт")],
        [KeyboardButton("💰 Рассчитать себестоимость")],
        [KeyboardButton("📖 Мои рецепты")],
        [KeyboardButton("⚖️ Пересчитать рецепт")],
        [KeyboardButton("📦 Остатки")],
        [KeyboardButton("🛒 Список покупок")],   # <-- новая кнопка
        [KeyboardButton("📊 Аналитика")],
        [KeyboardButton("📅 Заказы")],
        [KeyboardButton("❓ Помощь")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

async def show_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE, first_time=False):
    """Показывает главное меню"""
    if update.message:
        if first_time:
            user = update.effective_user
            name = user.first_name or "друг"
            greeting = f"👋 Привет, {name}! Я помогу управлять кондитерским производством.\nЧто будем делать?"
            await update.message.reply_text(greeting, reply_markup=get_main_keyboard())
        else:
            await update.message.reply_text("Главное меню:", reply_markup=get_main_keyboard())
    elif update.callback_query:
        await update.callback_query.message.reply_text("Главное меню:", reply_markup=get_main_keyboard())
async def handle_menu_buttons(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        text = update.message.text
        print(f"🔍 handle_menu_buttons получил текст: '{text}'")

        # ---- Главное меню ----
        if text == "📦 Ингредиенты":
            await update.message.reply_text("Выберите действие:", reply_markup=get_ingredients_submenu())
            return
        elif text == "🍰 Рецепты":
            await update.message.reply_text("Выберите действие:", reply_markup=get_recipes_submenu())
            return
        elif text == "💰 Продажи":
            await update.message.reply_text("Выберите действие:", reply_markup=get_sales_submenu())
            return
        elif text == "📊 Аналитика":
            await update.message.reply_text("Выберите действие:", reply_markup=get_analytics_submenu())
            return
        elif text == "🛒 Закупки":
            await update.message.reply_text("Выберите действие:", reply_markup=get_purchases_submenu())
            return
        elif text == "👥 Клиенты":
            await update.message.reply_text("Выберите действие:", reply_markup=get_customers_submenu())
            return
        elif text == "❓ Помощь":
            await help_command(update, context)
            return

        # ---- Подменю Ингредиенты ----
        elif text == "➕ Добавить ингредиент":
            # Запускаем диалог добавления ингредиента? Пока просто инструкция
            await update.message.reply_text(
                "Чтобы добавить ингредиент, отправьте:\n"
                "/add_ingredient название цена единица\n"
                "Или с ценой за упаковку: /add_ingredient название цена_упаковки вес_упаковки единица_веса\n\n"
                "Примеры:\n"
                "/add_ingredient мука 50 кг\n"
                "/add_ingredient масло 209.99 180 г"
            )
            return
        elif text == "📋 Список ингредиентов":
            await show_ingredients(update, context)
            return
        elif text == "🔄 Обновить цену":
            await update.message.reply_text(
                "Чтобы обновить цену, отправьте:\n"
                "/update_price название новая_цена\n"
                "Пример: /update_price мука 55"
            )
            return
        elif text == "📦 Закупка (партия)":
            await update.message.reply_text(
                "Чтобы зарегистрировать закупку, отправьте:\n"
                "/purchase ингредиент количество цена ГГГГ-ММ-ДД [поставщик]\n"
                "Пример: /purchase мука 10 500 2026-06-01 Мельница"
            )
            return
        elif text == "⏰ Сроки годности":
            await update.message.reply_text(
                "Чтобы установить сроки годности для ингредиента, отправьте:\n"
                "/set_shelf_life ингредиент общий_срок [срок_после_вскрытия]\n"
                "Пример: /set_shelf_life сливки 14 5\n\n"
                "Посмотреть истекающие сроки: /expiring"
            )
            return

        # ---- Подменю Рецепты ----
        elif text == "➕ Новый рецепт":
            await update.message.reply_text(
                "Чтобы добавить рецепт, отправьте:\n"
                "/add_recipe2 название тип базовое_количество: ингредиенты\n"
                "Пример: /add_recipe2 торт вес 1: мука 0.5, сахар 0.2, яйца 3"
            )
            return
        elif text == "📋 Мои рецепты":
            await list_recipes(update, context)
            return
        elif text == "🔍 Показать рецепт":
            await update.message.reply_text(
                "Введите название рецепта: /show_recipe название\n"
                "Пример: /show_recipe меренговый_рулет_белый"
            )
            return
        elif text == "⚖️ Пересчитать":
            await update.message.reply_text(
                "Чтобы пересчитать рецепт, отправьте:\n"
                "/scale название новое_количество [единица]\n"
                "Пример: /scale торт 2.5 кг"
            )
            return
        elif text == "💰 Себестоимость":
            await update.message.reply_text(
                "Введите название рецепта: /calculate название\n"
                "Пример: /calculate меренговый_рулет_белый"
            )
            return
        elif text == "📈 Прайс-лист":
            await price_list(update, context)
            return

                # ---- Подменю Продажи ----
        elif text == "📦 Новый заказ":
            await update.message.reply_text(
                "Введите название рецепта и количество:\n"
                "Например: `меренговый_рулет_фисташковый 2`\n\n"
                "Или используйте команду: /order название количество [цена]"
            )
            return
        elif text == "💵 Рекомендованная цена":
            await update.message.reply_text(
                "Введите название рецепта и количество:\n"
                "/price название количество\n"
                "Пример: /price меренговый_рулет_белый 2"
            )
            return
        elif text == "📊 Прибыль":
            await profit_command(update, context)  # вызываем новую команду
            return
        elif text == "🏆 Популярные":
            await popular(update, context)
            return
        elif text == "🔄 Возврат":
            await update.message.reply_text(
                "Чтобы оформить возврат, отправьте:\n"
                "/refund клиент рецепт количество [дата] [причина]\n"
                "Пример: /refund Иван меренговый_рулет_белый 1 не понравился"
            )
            return
        elif text == "❌ Списание":
            await update.message.reply_text(
                "Чтобы списать продукцию, отправьте:\n"
                "/write_off название количество [причина]\n"
                "Пример: /write_off мука 0.5 просыпалась"
            )
            return
        # ---- Подменю Аналитика ----
        elif text == "📈 Прибыль за месяц":
            await stats(update, context)  # пока то же, что и stats
            return
        elif text == "📉 Самые прибыльные":
            await popular(update, context)  # пока то же, позже сделаем /top_profit
            return
        elif text == "📊 Отчёт Excel":
            await report_xlsx(update, context)
            return

        # ---- Подменю Закупки ----
        elif text == "🛒 Список покупок":
            await shopping(update, context)
            return
        elif text == "📦 Запланировать":
            await update.message.reply_text(
                "Чтобы запланировать производство, отправьте:\n"
                "/plan рецепт количество ГГГГ-ММ-ДД\n"
                "Пример: /plan меренговый_рулет_белый 5 2026-03-20"
            )
            return
        elif text == "⏳ Истекающие сроки":
            await expiring(update, context)
            return

        # ---- Подменю Клиенты ----
        elif text == "➕ Новый клиент":
            await update.message.reply_text(
                "Чтобы добавить клиента, отправьте:\n"
                "/add_customer имя телефон [адрес]\n"
                "Пример: /add_customer Иван +79991234567"
            )
            return
        elif text == "📅 Создать заказ":
            await update.message.reply_text(
                "Чтобы создать предзаказ, отправьте:\n"
                "/order клиент рецепт количество ГГГГ-ММ-ДД\n"
                "Пример: /order Иван меренговый_рулет_белый 2 2026-03-20"
            )
            return
        elif text == "📋 Заказы на дату":
            await update.message.reply_text(
                "Введите дату: /orders ГГГГ-ММ-ДД\n"
                "Пример: /orders 2026-03-20"
            )
            return
        elif text == "🔔 Напоминания":
            await remind_orders(update, context)
            return

        # ---- Кнопка "Назад" ----
        elif text == "« Назад":
            await update.message.reply_text("Главное меню:", reply_markup=get_main_keyboard())
            return

        else:
            print(f"⚠️ Неизвестный текст: {text}")
            # Если текст не распознан, можно просто показать главное меню
            # await update.message.reply_text("Не понял команду. Выберите действие:", reply_markup=get_main_keyboard())
            return

    except Exception as e:
        print(f"❌ Ошибка в handle_menu_buttons: {e}")
        import traceback
        traceback.print_exc()
        await update.message.reply_text("Произошла внутренняя ошибка. Пожалуйста, сообщите разработчику.")
async def show_stock(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not ingredients:
        await update.message.reply_text("Список ингредиентов пуст")
        return
    msg = "📦 Текущие остатки:\n"
    for name in sorted(ingredients.keys()):
        data = ingredients[name]
        stock = data.get('stock', 0.0)
        unit = data['unit']
        line = f"• {name}: {stock:.2f} {unit}\n"
        if len(msg) + len(line) > 4000:
            await update.message.reply_text(msg)
            msg = "📦 Продолжение остатков:\n"
        msg += line
    await update.message.reply_text(msg)
async def use_recipe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Списать ингредиенты на приготовление: /use рецепт [количество] [фактическая_цена]"""
    args = context.args
    if len(args) < 1:
        await update.message.reply_text(
            "Формат: /use название_рецепта [количество] [фактическая_цена]\n"
            "Примеры:\n"
            "/use меренговый_рулет_белый\n"
            "/use меренговый_рулет_белый 2\n"
            "/use меренговый_рулет_белый 2 1500"
        )
        return

    # Определяем количество и цену
    possible_qty = None
    possible_price = None
    name_parts = args[:]

    if len(args) >= 2:
        try:
            float(args[-1].replace(',', '.'))
            if len(args) >= 3:
                try:
                    float(args[-2].replace(',', '.'))
                    possible_qty = float(args[-2].replace(',', '.'))
                    possible_price = float(args[-1].replace(',', '.'))
                    name_parts = args[:-2]
                except ValueError:
                    possible_qty = float(args[-1].replace(',', '.'))
                    name_parts = args[:-1]
            else:
                possible_qty = float(args[-1].replace(',', '.'))
                name_parts = args[:-1]
        except ValueError:
            pass

    name = ' '.join(name_parts).lower()
    qty = possible_qty if possible_qty is not None else 1.0
    price = possible_price

    result_msg = await execute_sale(update, context, name, qty, price)
    await update.message.reply_text(result_msg)
async def execute_sale(update: Update, context: ContextTypes.DEFAULT_TYPE, recipe_name: str, qty: float, price: float = None):
    """
    Выполняет продажу: списывает ингредиенты, записывает продажу.
    Возвращает сообщение для пользователя.
    """
    if recipe_name not in recipes:
        return f"❌ Рецепт '{recipe_name}' не найден."

    data = recipes[recipe_name]

    # Получаем ингредиенты рецепта
    if isinstance(data, dict) and "ingredients" in data:
        ing_dict = data["ingredients"]
    else:
        ing_dict = data

    # Масштабируем на нужное количество
    needed = {ing: ing_qty * qty for ing, ing_qty in ing_dict.items()}

    # Проверяем наличие по партиям
    missing = []
    total_ing_cost = 0.0

    for ing_name, need_qty in needed.items():
        if ing_name not in ingredients:
            missing.append(ing_name)
            continue

        # Находим все активные партии для этого ингредиента
        ing_batches = [b for b in batches if b.get('is_active', True) and b['ingredient'] == ing_name]
        ing_batches.sort(key=lambda b: b['expiry_date'])

        remaining = need_qty
        for batch in ing_batches:
            if remaining <= 0:
                break

            # Определяем актуальную дату истечения
            if batch.get('opened_date'):
                expiry = datetime.fromisoformat(batch['expiry_after_open']).date()
            else:
                expiry = datetime.fromisoformat(batch['expiry_date']).date()

            # Проверяем, не просрочена ли партия
            if expiry < datetime.now().date():
                continue

            # Если партия не открыта, открываем её при первом списании
            if not batch.get('opened_date'):
                batch['opened_date'] = datetime.now().date().isoformat()
                after_open_days = ingredients[ing_name].get('shelf_life_after_open', ingredients[ing_name].get('shelf_life', 30))
                batch['expiry_after_open'] = (datetime.now().date() + timedelta(days=after_open_days)).isoformat()

            available = batch['current_quantity']
            take = min(available, remaining)
            if take <= 0:
                continue

            # Считаем себестоимость этой части
            unit_cost = batch['cost'] / batch['initial_quantity']
            cost_part = take * unit_cost
            total_ing_cost += cost_part

            # Уменьшаем остаток в партии
            batch['current_quantity'] -= take
            remaining -= take
            if batch['current_quantity'] <= 0:
                batch['is_active'] = False

        if remaining > 0:
            missing.append(f"{ing_name} (не хватает {remaining:.2f} {ingredients[ing_name]['unit']})")

    if missing:
        return f"❌ Недостаточно ингредиентов:\n" + "\n".join(missing)

    # Сохраняем изменения в партиях
    save_batches()

    # Пересчитываем общий остаток для затронутых ингредиентов
    for ing_name in needed.keys():
        recalc_ingredient_stock(ing_name)

    # Дополнительные расходы
    packaging = data.get('packaging', 0.0) if isinstance(data, dict) else 0.0
    work_hours = data.get('work_hours', 0.0) if isinstance(data, dict) else 0.0
    hourly_rate = settings.get('hourly_rate', 0.0)
    work_cost = work_hours * hourly_rate * qty if work_hours and hourly_rate else 0.0
    markup = data.get('markup') if isinstance(data, dict) else None

    total_cost = total_ing_cost + packaging * qty + work_cost

    # Если цена не передана, рассчитываем по наценке
    if price is None:
        if markup is not None:
            price = total_cost * (1 + markup / 100)
        else:
            price = None

    if price is not None:
        profit = price - total_cost
    else:
        profit = None

    # Запись о продаже
    sale_record = {
        "date": datetime.now().isoformat(),
        "recipe": recipe_name,
        "quantity": qty,
        "cost": total_ing_cost,
        "cost_with_extras": total_cost,
        "price": price,
        "profit": profit
    }
    sales.append(sale_record)
    save_sales()

    # Формируем ответ
    msg = f"✅ Приготовлено {qty} шт '{recipe_name}'. Ингредиенты списаны.\n"
    msg += f"💰 Себестоимость ингредиентов: {total_ing_cost:.2f} руб\n"
    if packaging or work_cost:
        msg += f"🧾 Полная себестоимость: {total_cost:.2f} руб\n"
    if price is not None:
        msg += f"💵 Цена продажи: {price:.2f} руб\n"
        if profit is not None:
            msg += f"💸 Прибыль: {profit:.2f} руб\n"
            if total_cost > 0:
                msg += f"📊 Рентабельность: {(profit/total_cost*100):.1f}%"
    else:
        msg += f"❓ Наценка не задана, цена продажи не рассчитана."

    return msg
def recalc_ingredient_stock(ingredient):
    """Пересчитывает общий остаток ингредиента по всем активным партиям и обновляет ingredients.json"""
    total = sum(b['current_quantity'] for b in batches if b.get('is_active', True) and b['ingredient'] == ingredient)
    if ingredient in ingredients:
        ingredients[ingredient]['stock'] = total
    else:
        # если ингредиента почему-то нет в ingredients (маловероятно), создадим
        ingredients[ingredient] = {'stock': total, 'unit': 'кг'}  # unit нужно взять из партии
        # но лучше не создавать, а сообщить об ошибке
        logger.error(f"Ингредиент {ingredient} отсутствует в ingredients, но есть партии")
    save_data(ingredients, INGREDIENTS_FILE)
async def write_off(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Списать ингредиенты или готовую продукцию: /write_off <название> <количество> [причина]"""
    if len(context.args) < 2:
        await update.message.reply_text(
            "Формат: /write_off название количество [причина]\n"
            "Примеры:\n"
            "/write_off мука 0.5 просыпалась\n"
            "/write_off меренговый_рулет_белый 2 брак"
        )
        return

    # Пытаемся выделить название (может быть из нескольких слов) и количество
    *name_parts, qty_str = context.args[:2]  # первые два аргумента точно есть
    reason = ' '.join(context.args[2:]) if len(context.args) > 2 else "не указана"
    name = ' '.join(name_parts).lower()
    try:
        qty = float(qty_str.replace(',', '.'))
    except ValueError:
        await update.message.reply_text("Ошибка! Количество должно быть числом")
        return

    # Проверяем, ингредиент это или рецепт
    if name in ingredients:
        # Списание ингредиента со склада
        if 'stock' not in ingredients[name]:
            ingredients[name]['stock'] = 0.0
        if ingredients[name]['stock'] < qty:
            await update.message.reply_text(f"❌ Недостаточно '{name}' на складе. Есть {ingredients[name]['stock']} {ingredients[name]['unit']}")
            return
        ingredients[name]['stock'] -= qty
        save_data(ingredients, INGREDIENTS_FILE)
        record = {
            "date": datetime.now().isoformat(),
            "type": "write_off",
            "item": name,
            "quantity": qty,
            "unit": ingredients[name]['unit'],
            "reason": reason
        }
        msg = f"✅ Списано {qty} {ingredients[name]['unit']} '{name}'. Причина: {reason}"
    elif name in recipes:
        # Списание готовой продукции – списываем ингредиенты рецепта
        data = recipes[name]
        if isinstance(data, dict) and "ingredients" in data:
            ing_dict = data["ingredients"]
        else:
            ing_dict = data
        # Проверяем наличие всех ингредиентов
        missing = []
        for ing_name, ing_qty in ing_dict.items():
            need = ing_qty * qty
            if ing_name not in ingredients:
                missing.append(ing_name)
            else:
                stock = ingredients[ing_name].get('stock', 0.0)
                if stock < need:
                    missing.append(f"{ing_name} (нужно {need}, есть {stock})")
        if missing:
            await update.message.reply_text(f"❌ Недостаточно ингредиентов для списания {qty} шт '{name}':\n" + "\n".join(missing))
            return
        # Списание
        for ing_name, ing_qty in ing_dict.items():
            need = ing_qty * qty
            ingredients[ing_name]['stock'] -= need
        save_data(ingredients, INGREDIENTS_FILE)
        record = {
            "date": datetime.now().isoformat(),
            "type": "write_off",
            "item": name,
            "quantity": qty,
            "unit": "шт",
            "reason": reason
        }
        msg = f"✅ Списано {qty} шт '{name}'. Причина: {reason}"
    else:
        await update.message.reply_text(f"Ингредиент или рецепт '{name}' не найден")
        return

    writeoffs.append(record)
    save_writeoffs()
    await update.message.reply_text(msg)
async def refund(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Оформить возврат товара от клиента: /refund клиент рецепт количество [дата] [причина]"""
    if len(context.args) < 3:
        await update.message.reply_text(
            "Формат: /refund клиент рецепт количество [дата] [причина]\n"
            "Пример: /refund Иван меренговый_рулет_белый 1 2026-03-15 не понравился"
        )
        return

    customer = context.args[0]
    recipe_name = context.args[1].lower()
    qty_str = context.args[2]
    try:
        qty = float(qty_str.replace(',', '.'))
    except ValueError:
        await update.message.reply_text("Ошибка! Количество должно быть числом")
        return

    reason = "не указана"
    refund_date = datetime.now().date()  # дата возврата
    if len(context.args) >= 4:
        try:
            # пытаемся распарсить дату
            refund_date = datetime.strptime(context.args[3], "%Y-%m-%d").date()
            if len(context.args) > 4:
                reason = ' '.join(context.args[4:])
        except ValueError:
            # если не дата, значит это причина
            reason = ' '.join(context.args[3:])

    if customer not in customers:
        await update.message.reply_text(f"Клиент '{customer}' не найден")
        return

    if recipe_name not in recipes:
        await update.message.reply_text(f"Рецепт '{recipe_name}' не найден")
        return

    # Запись о возврате (без изменения остатков)
    record = {
        "date": datetime.now().isoformat(),
        "type": "refund",
        "customer": customer,
        "recipe": recipe_name,
        "quantity": qty,
        "reason": reason,
        "refund_date": refund_date.isoformat()
    }
    writeoffs.append(record)
    save_writeoffs()

    await update.message.reply_text(
        f"✅ Оформлен возврат {qty} шт '{recipe_name}' от клиента {customer}. "
        f"Ингредиенты не восстанавливаются (убыток). Причина: {reason}"
    )
async def price_history_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показать историю изменения цены ингредиента: /price_history <название>"""
    if not context.args:
        await update.message.reply_text("Укажите название ингредиента: /price_history мука")
        return
    name = ' '.join(context.args).lower()
    records = [r for r in price_history if r.get('ingredient') == name]
    if not records:
        await update.message.reply_text(f"Для ингредиента '{name}' нет истории изменений.")
        return
    records.sort(key=lambda x: x['date'])
    msg = f"📈 История изменения цены для '{name}':\n\n"
    for r in records:
        date_str = datetime.fromisoformat(r['date']).strftime("%d.%m.%Y %H:%M")
        unit = r.get('unit', 'кг')  # если единица не сохранилась, по умолчанию кг
        if r['old_price'] is None:
            msg += f"🟢 {date_str}: добавлен, цена = {r['new_price']:.2f} руб/{unit}\n"
        else:
            msg += f"🔄 {date_str}: {r['old_price']:.2f} → {r['new_price']:.2f} руб/{unit}\n"
    if len(msg) > 4000:
        parts = [msg[i:i+4000] for i in range(0, len(msg), 4000)]
        for part in parts:
            await update.message.reply_text(part)
    else:
        await update.message.reply_text(msg)
async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Статистика продаж: /stats [день|неделя|месяц|год]"""
    try:
        period = context.args[0].lower() if context.args else "месяц"
        now = datetime.now()
        if period == "день":
            start = now - timedelta(days=1)
        elif period == "неделя":
            start = now - timedelta(weeks=1)
        elif period == "месяц":
            start = now - timedelta(days=30)
        elif period == "год":
            start = now - timedelta(days=365)
        else:
            await update.message.reply_text("Период может быть: день, неделя, месяц, год")
            return

        total_revenue = 0.0
        total_cost = 0.0
        total_profit = 0.0
        count = 0

        for sale in sales:
            sale_date = datetime.fromisoformat(sale['date'])
            if sale_date >= start:
                total_revenue += sale.get('price', 0.0) or 0.0
                total_cost += sale.get('cost_with_extras', sale['cost'])
                total_profit += sale.get('profit', 0.0) or 0.0
                count += 1

        if count == 0:
            await update.message.reply_text(f"Нет продаж за {period}.")
            return

        msg = f"📊 *Статистика за {period}:*\n"
        msg += f"• Продано десертов: {count}\n"
        msg += f"• Выручка: {total_revenue:.2f} руб\n"
        msg += f"• Себестоимость: {total_cost:.2f} руб\n"
        msg += f"• Прибыль: {total_profit:.2f} руб\n"
        if total_cost > 0:
            msg += f"• Рентабельность: {(total_profit/total_cost*100):.1f}%"
        await update.message.reply_text(msg)
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка в stats: {e}")
async def low_stock(update: Update, context: ContextTypes.DEFAULT_TYPE):
    threshold = 1.0
    if context.args:
        try:
            threshold = float(context.args[0].replace(',', '.'))
        except:
            pass
    if not ingredients:
        await update.message.reply_text("Список ингредиентов пуст")
        return
    low = []
    for name, data in ingredients.items():
        stock = data.get('stock', 0.0)
        if stock < threshold:
            low.append(f"• {name}: {stock:.2f} {data['unit']}")
    if low:
        msg = f"⚠️ *Ингредиенты с остатком менее {threshold}:*\n" + "\n".join(low)
    else:
        msg = f"✅ Все остатки выше {threshold}."
    await update.message.reply_text(msg)
async def popular(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Топ-5 рецептов по продажам"""
    try:
        recipe_count = {}
        for sale in sales:
            recipe = sale['recipe']
            recipe_count[recipe] = recipe_count.get(recipe, 0) + sale.get('quantity', 1)

        if not recipe_count:
            await update.message.reply_text("Пока нет продаж.")
            return

        sorted_recipes = sorted(recipe_count.items(), key=lambda x: x[1], reverse=True)[:5]
        msg = "🏆 *Топ-5 рецептов:*\n"
        for i, (recipe, cnt) in enumerate(sorted_recipes, 1):
            msg += f"{i}. {recipe}: {cnt} шт\n"
        await update.message.reply_text(msg)
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка в popular: {e}")

async def list_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показать заказы на указанную дату (по умолчанию сегодня): /orders [ГГГГ-ММ-ДД]"""
    # Парсим дату
    if context.args:
        try:
            target_date = datetime.strptime(context.args[0], "%Y-%m-%d").date()
        except ValueError:
            await update.message.reply_text("Неверный формат даты. Используйте ГГГГ-ММ-ДД")
            return
    else:
        target_date = datetime.now().date()

    msg_lines = [f"📅 Заказы на {target_date}:"]
    found = False

    for order in orders:
        # Определяем дату заказа
        order_date = None
        if 'due_date' in order:
            try:
                order_date = datetime.fromisoformat(order['due_date']).date()
            except (ValueError, TypeError):
                continue
        elif 'date' in order:
            try:
                # Предполагаем, что 'date' может быть в формате YYYY-MM-DD или полный ISO
                date_str = order['date']
                if 'T' in date_str:
                    date_str = date_str.split('T')[0]
                order_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except (ValueError, TypeError, KeyError):
                continue
        else:
            continue  # нет даты – пропускаем

        if order_date != target_date:
            continue

        found = True

        # Извлекаем общие поля
        recipe = order.get('recipe', 'Неизвестный рецепт')
        quantity = order.get('quantity', 0)

        # Определяем тип заказа и формируем строку
        if 'customer' in order:
            # Это предзаказ
            customer = order.get('customer', 'Неизвестный клиент')
            status = order.get('status', 'неизвестен')
            line = f"• {customer}: {recipe} – {quantity} шт, статус: {status}"
        else:
            # Обычный заказ (без клиента)
            # Можно показать стоимость или прибыль, если есть
            cost = order.get('cost', 0)
            revenue = order.get('revenue', order.get('price', 0))
            profit = order.get('profit', 0)
            line = f"• {recipe} – {quantity} шт"
            if revenue:
                line += f", выручка: {revenue:.0f} ₽"
            if profit:
                line += f", прибыль: {profit:.0f} ₽"

        msg_lines.append(line)

    if not found:
        msg_lines.append("Нет заказов на эту дату.")

    msg = "\n".join(msg_lines)
    await update.message.reply_text(msg)
async def add_customer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Добавить клиента: /add_customer имя телефон [адрес]"""
    if len(context.args) < 2:
        await update.message.reply_text(
            "Формат: /add_customer имя телефон [адрес]\n"
            "Пример: /add_customer Анна +79991234567"
        )
        return
    name = context.args[0]
    phone = context.args[1]
    address = ' '.join(context.args[2:]) if len(context.args) > 2 else ""
    if name in customers:
        await update.message.reply_text(f"Клиент с именем '{name}' уже существует. Используйте другое имя или удалите старого.")
        return
    customers[name] = {"phone": phone, "address": address, "notes": ""}
    save_customers()
    await update.message.reply_text(f"✅ Клиент {name} добавлен")
async def create_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Создать предзаказ: /order клиент рецепт количество ГГГГ-ММ-ДД"""
    if len(context.args) < 4:
        await update.message.reply_text(
            "Формат: /order клиент рецепт количество ГГГГ-ММ-ДД\n"
            "Пример: /order Анна меренговый_рулет_белый 2 2026-03-20"
        )
        return
    customer = context.args[0]
    recipe_name = context.args[1].lower()
    qty_str = context.args[2]
    date_str = context.args[3]
    try:
        qty = float(qty_str.replace(',', '.'))
        date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
        await update.message.reply_text("Ошибка! Количество должно быть числом, дата в формате ГГГГ-ММ-ДД")
        return
    if customer not in customers:
        await update.message.reply_text(f"Клиент '{customer}' не найден. Сначала добавьте через /add_customer")
        return
    if recipe_name not in recipes:
        await update.message.reply_text(f"Рецепт '{recipe_name}' не найден")
        return
    order = {
        "customer": customer,
        "recipe": recipe_name,
        "quantity": qty,
        "due_date": date.isoformat(),
        "status": "pending"
    }
    orders.append(order)
    save_orders()
    await update.message.reply_text(f"✅ Заказ для {customer} на {qty} шт '{recipe_name}' к {date_str} создан.")
async def remind_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Напомнить о заказах на сегодня и завтра: /remind"""
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)
    msg = "🔔 *Напоминание о заказах:*\n"
    found = False
    for order in orders:
        order_date = datetime.fromisoformat(order['due_date']).date()
        if order_date == today:
            msg += f"• СЕГОДНЯ: {order['customer']} – {order['recipe']} ({order['quantity']} шт)\n"
            found = True
        elif order_date == tomorrow:
            msg += f"• ЗАВТРА: {order['customer']} – {order['recipe']} ({order['quantity']} шт)\n"
            found = True
    if not found:
        msg = "Нет заказов на сегодня и завтра."
    await update.message.reply_text(msg)
async def set_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Установить текущий чат как административный для напоминаний"""
    chat_id = update.effective_chat.id
    settings['admin_chat_id'] = chat_id
    save_settings()
    await update.message.reply_text(f"✅ Административный чат установлен (ID: {chat_id})")
async def menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await show_main_menu(update, context, first_time=False)
async def set_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Установить категорию для рецепта: /set_category рецепт категория"""
    if len(context.args) < 2:
        await update.message.reply_text(
            "Формат: /set_category рецепт категория\n"
            "Пример: /set_category меренговый_рулет_белый рулеты"
        )
        return
    # Название рецепта может состоять из нескольких слов, поэтому все аргументы кроме последнего — это название
    *name_parts, category = context.args
    name = ' '.join(name_parts).lower()
    if name not in recipes:
        await update.message.reply_text(f"Рецепт '{name}' не найден")
        return
    if isinstance(recipes[name], dict):
        recipes[name]['category'] = category.lower()
    else:
        # Если рецепт в старом формате (не словарь), преобразуем
        recipes[name] = {"ingredients": recipes[name], "category": category.lower()}
    save_data(recipes, RECIPES_FILE)
    await update.message.reply_text(f"✅ Категория '{category}' установлена для рецепта '{name}'")
# ---------- Ингредиенты ----------
async def add_ingredient(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        args = context.args
        if len(args) < 3:
            await update.message.reply_text(
               "Чтобы добавить ингредиент, отправьте:\n"
               "/add_ingredient название цена единица\n"
               "или с ценой за упаковку: /add_ingredient название цена_упаковки вес_упаковки единица_веса\n\n"
               "Примеры:\n"
               "/add_ingredient мука 50 кг\n"
               "/add_ingredient масло 209.99 180 г"
            )
            return

        if len(args) >= 4:
            # Новый формат: цена, количество, единица
            *name_parts, price_str, qty_str, unit = args
            name = ' '.join(name_parts).lower()
            price_pack = float(price_str.replace(',', '.'))
            qty_pack = float(qty_str.replace(',', '.'))
            unit = unit.lower()

            allowed_units = ['кг','г','л','мл','шт','kg','g','l','ml','pcs']
            if unit not in allowed_units:
                await update.message.reply_text(f"Единица должна быть одной из: {', '.join(allowed_units)}")
                return

            # Пересчёт в базовую единицу (кг, л, шт)
            if unit in ['г', 'кг']:
                if unit == 'г':
                    qty_kg = qty_pack / 1000.0
                else:
                    qty_kg = qty_pack
                price_per_kg = price_pack / qty_kg
                price_per_kg = round(price_per_kg, 2)
                ingredients[name] = {"price": price_per_kg, "unit": "кг"}
                save_data(ingredients, INGREDIENTS_FILE)
                record_price_history(name, None, price_per_kg, "кг")
                await update.message.reply_text(
                    f"✅ Ингредиент '{name}' добавлен: {price_per_kg} руб/кг "
                    f"(рассчитано из {price_pack} руб за {qty_pack} {unit})"
                )
            elif unit in ['мл', 'л']:
                if unit == 'мл':
                    qty_l = qty_pack / 1000.0
                else:
                    qty_l = qty_pack
                price_per_l = price_pack / qty_l
                price_per_l = round(price_per_l, 2)
                ingredients[name] = {"price": price_per_l, "unit": "л"}
                save_data(ingredients, INGREDIENTS_FILE)
                record_price_history(name, None, price_per_l, "л")
                await update.message.reply_text(
                    f"✅ Ингредиент '{name}' добавлен: {price_per_l} руб/л "
                    f"(рассчитано из {price_pack} руб за {qty_pack} {unit})"
                )
            elif unit == 'шт':
                price_per_pcs = price_pack / qty_pack
                price_per_pcs = round(price_per_pcs, 2)
                ingredients[name] = {"price": price_per_pcs, "unit": "шт"}
                save_data(ingredients, INGREDIENTS_FILE)
                record_price_history(name, None, price_per_pcs, "шт")
                await update.message.reply_text(
                    f"✅ Ингредиент '{name}' добавлен: {price_per_pcs} руб/шт "
                    f"(рассчитано из {price_pack} руб за {qty_pack} шт)"
                )
        else:
            # Старый формат: цена и единица
            *name_parts, price_str, unit = args
            name = ' '.join(name_parts).lower()
            price = float(price_str.replace(',', '.'))
            unit = unit.lower()
            allowed_units = ['кг', 'г', 'шт', 'л', 'мл']
            if unit not in allowed_units:
                await update.message.reply_text(f"Ошибка! Единица должна быть одной из: {', '.join(allowed_units)}")
                return
            ingredients[name] = {"price": price, "unit": unit}
            save_data(ingredients, INGREDIENTS_FILE)
            record_price_history(name, None, price, unit)
            await update.message.reply_text(f"✅ Ингредиент '{name}' добавлен: {price} руб/{unit}")
    except ValueError:
        await update.message.reply_text("Ошибка! Цена и количество должны быть числами (например, 50 или 45.5)")
    except Exception as e:
        await update.message.reply_text(f"Произошла ошибка: {e}")
async def show_ingredients(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.callback_query:
        await update.callback_query.answer()
        reply_func = update.callback_query.edit_message_text
    else:
        reply_func = update.message.reply_text
    if not ingredients:
        await reply_func("Список ингредиентов пуст")
        return
    message = "📋 *Список ингредиентов (по алфавиту):*\n"
    for name in sorted(ingredients.keys()):
        data = ingredients[name]
        message += f"• {name}: {data['price']} руб/{data['unit']}\n"
    await reply_func(message)

async def remove_ingredient(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text(
        "Укажите название ингредиента: /remove_ingredient название\n"
        "Пример: /remove_ingredient мука"
        )
        return
    name = ' '.join(context.args).lower()
    if name in ingredients:
        del ingredients[name]
        save_data(ingredients, INGREDIENTS_FILE)
        await update.message.reply_text(f"✅ Ингредиент '{name}' удалён")
    else:
        await update.message.reply_text(f"Ингредиент '{name}' не найден")

async def update_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 2:
        await update.message.reply_text(
            "Формат: /update_price название новая_цена\n"
            "Пример: /update_price мука 55"
        )
        return
    *name_parts, price_str = context.args
    name = ' '.join(name_parts).lower()
    try:
        new_price = float(price_str.replace(',', '.'))
        if name in ingredients:
            old_price = ingredients[name]["price"]
            ingredients[name]["price"] = new_price
            save_data(ingredients, INGREDIENTS_FILE)
            record_price_history(name, old_price, new_price, ingredients[name]["unit"])
            await update.message.reply_text(f"✅ Цена '{name}' обновлена: {new_price} руб/{ingredients[name]['unit']}")
        else:
            await update.message.reply_text(f"Ингредиент '{name}' не найден")
    except ValueError:
        await update.message.reply_text("Ошибка! Цена должна быть числом")
# ---------- Рецепты ----------
async def add_recipe(update: Update, context: ContextTypes.DEFAULT_TYPE):  # старый формат
    try:
        text = update.message.text.replace('/add_recipe', '', 1).strip()
        if ':' not in text:
            await update.message.reply_text(
                "Ошибка! Нужно: /add_recipe Название: порции; ингредиент количество, ...\n"
                "Пример: /add_recipe Омлет: 2; яйца 3, молоко 0.1"
            )
            return
        name_part, rest = text.split(':', 1)
        name = name_part.strip().lower()
        if ';' in rest:
            portions_part, ingredients_part = rest.split(';', 1)
            try:
                portions = float(portions_part.strip())
            except ValueError:
                await update.message.reply_text("Ошибка! Количество порций должно быть числом")
                return
        else:
            portions = 1
            ingredients_part = rest
        recipe_ingredients = {}
        for item in ingredients_part.split(','):
            item = item.strip()
            if not item:
                continue
            parts = item.rsplit(' ', 1)
            if len(parts) != 2:
                await update.message.reply_text(f"Ошибка в элементе: '{item}'. Должно быть 'название количество'")
                return
            ing_name = parts[0].strip().lower()
            try:
                qty = float(parts[1].replace(',', '.'))
            except ValueError:
                await update.message.reply_text(f"Ошибка: количество должно быть числом в '{item}'")
                return
            recipe_ingredients[ing_name] = qty
        recipes[name] = {"ingredients": recipe_ingredients, "portions": portions}
        save_data(recipes, RECIPES_FILE)
        await update.message.reply_text(f"✅ Рецепт '{name}' добавлен! ({portions} порций)")
    except Exception as e:
        await update.message.reply_text(f"Произошла ошибка: {e}")

async def add_recipe_scaled(update: Update, context: ContextTypes.DEFAULT_TYPE):  # масштабируемый формат
    try:
        text = update.message.text.replace('/add_recipe2', '', 1).strip()
        if ':' not in text:
            await update.message.reply_text(
                "Ошибка! Нужно: /add_recipe2 название тип базовое_количество: ингредиенты\n"
                "Пример: /add_recipe2 торт вес 1: мука 0.5, сахар 0.2"
            )
            return
        left, right = text.split(':', 1)
        parts = left.strip().split()
        if len(parts) < 3:
            await update.message.reply_text("Укажите название, тип (вес/штук) и базовое количество")
            return
        name = parts[0].lower()
        type_str = parts[1].lower()
        try:
            base_qty = float(parts[2].replace(',', '.'))
        except ValueError:
            await update.message.reply_text("Ошибка! Базовое количество должно быть числом")
            return
        if type_str == 'вес':
            recipe_type = 'weight'
            base_unit = 'кг'
        elif type_str == 'штук':
            recipe_type = 'pcs'
            base_unit = 'шт'
        else:
            await update.message.reply_text("Ошибка! Тип должен быть 'вес' или 'штук'")
            return
        ingredients_list = right.strip().split(',')
        recipe_ingredients = {}
        for item in ingredients_list:
            item = item.strip()
            if not item:
                continue
            parts = item.rsplit(' ', 1)
            if len(parts) != 2:
                await update.message.reply_text(f"Ошибка в элементе: '{item}'. Должно быть 'название количество'")
                return
            ing_name = parts[0].strip().lower()
            try:
                qty = float(parts[1].replace(',', '.'))
            except ValueError:
                await update.message.reply_text(f"Ошибка: количество должно быть числом в '{item}'")
                return
            recipe_ingredients[ing_name] = qty
        recipes[name] = {"type": recipe_type, "base_qty": base_qty, "ingredients": recipe_ingredients}
        save_data(recipes, RECIPES_FILE)
        await update.message.reply_text(f"✅ Рецепт '{name}' добавлен! (тип: {type_str}, база {base_qty} {base_unit})")
    except Exception as e:
        await update.message.reply_text(f"Произошла ошибка: {e}")

async def list_recipes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показать список рецептов. Если указана категория, показать только из неё."""
    try:
        category = None
        if context.args:
            category = ' '.join(context.args).lower()
        
        if update.callback_query:
            await update.callback_query.answer()
            reply_func = update.callback_query.edit_message_text
        else:
            reply_func = update.message.reply_text

        if not recipes:
            await reply_func("Список рецептов пуст")
            return

        # Фильтруем рецепты по категории, если она указана
        filtered_recipes = {}
        for name, data in recipes.items():
            if category:
                cat = data.get('category') if isinstance(data, dict) else None
                if cat and cat == category:
                    filtered_recipes[name] = data
            else:
                filtered_recipes[name] = data

        if not filtered_recipes:
            await reply_func(f"Нет рецептов в категории '{category}'." if category else "Список рецептов пуст")
            return

        # Формируем заголовок
        header = f"📖 Список рецептов{f' в категории {category}' if category else ''}:\n"
        
        # Функция для отправки длинных сообщений частями
        async def send_long_message(initial_msg, parts_generator):
            msg = initial_msg
            for line in parts_generator:
                if len(msg) + len(line) > 4000:
                    await reply_func(msg)
                    msg = "📖 Продолжение списка:\n"
                msg += line
            if msg:
                await reply_func(msg)
        
        # Генератор строк для каждого рецепта
        def generate_lines():
            for name, data in filtered_recipes.items():
                if isinstance(data, dict) and "type" in data:
                    rtype = "весовой" if data["type"] == "weight" else "штучный"
                    base = data["base_qty"]
                    unit = "кг" if data["type"] == "weight" else "шт"
                    ing_list = [f"{ing} {qty} {ingredients.get(ing, {}).get('unit', 'ед')}" for ing, qty in data["ingredients"].items()]
                    line = f"• {name} ({rtype}, база {base} {unit}): {', '.join(ing_list)}\n"
                else:
                    if isinstance(data, dict) and "ingredients" in data:
                        ing_dict = data["ingredients"]
                        portions = data.get("portions", 1)
                    else:
                        ing_dict = data
                        portions = 1
                    ing_list = [f"{ing} {qty} {ingredients.get(ing, {}).get('unit', 'ед')}" for ing, qty in ing_dict.items()]
                    line = f"• {name} ({portions} порц.): {', '.join(ing_list)}\n"
                yield line
        
        await send_long_message(header, generate_lines())
        
    except Exception as e:
        error_msg = f"❌ Ошибка в команде /recipes: {e}"
        await update.message.reply_text(error_msg)
async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Статистика продаж: /stats [день|неделя|месяц|год]"""
    period = context.args[0].lower() if context.args else "месяц"
    now = datetime.now()
    if period == "день":
        start = now - timedelta(days=1)
    elif period == "неделя":
        start = now - timedelta(weeks=1)
    elif period == "месяц":
        start = now - timedelta(days=30)
    elif period == "год":
        start = now - timedelta(days=365)
    else:
        await update.message.reply_text("Период может быть: день, неделя, месяц, год")
        return

    total_revenue = 0.0
    total_cost = 0.0
    total_profit = 0.0
    count = 0

    for sale in sales:
        sale_date = datetime.fromisoformat(sale['date'])
        if sale_date >= start:
            total_revenue += sale.get('price', 0.0) or 0.0
            total_cost += sale.get('cost_with_extras', sale['cost'])
            total_profit += sale.get('profit', 0.0) or 0.0
            count += 1

    if count == 0:
        await update.message.reply_text(f"Нет продаж за {period}.")
        return

    msg = f"📊 *Статистика за {period}:*\n"
    msg += f"• Продано десертов: {count}\n"
    msg += f"• Выручка: {total_revenue:.2f} руб\n"
    msg += f"• Себестоимость: {total_cost:.2f} руб\n"
    msg += f"• Прибыль: {total_profit:.2f} руб\n"
    if total_cost > 0:
        msg += f"• Рентабельность: {(total_profit/total_cost*100):.1f}%"
    await update.message.reply_text(msg)
async def popular(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Топ-5 рецептов по продажам"""
    recipe_count = {}
    for sale in sales:
        recipe = sale['recipe']
        recipe_count[recipe] = recipe_count.get(recipe, 0) + sale.get('quantity', 1)

    if not recipe_count:
        await update.message.reply_text("Пока нет продаж.")
        return

    sorted_recipes = sorted(recipe_count.items(), key=lambda x: x[1], reverse=True)[:5]
    msg = "🏆 *Топ-5 рецептов:*\n"
    for i, (recipe, cnt) in enumerate(sorted_recipes, 1):
        msg += f"{i}. {recipe}: {cnt} шт\n"
    await update.message.reply_text(msg)
async def remove_recipe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text(
        "Укажите название рецепта: /remove_recipe название\n"
        "Пример: /remove_recipe омлет"
        )
        return
    name = ' '.join(context.args).strip().lower()
    if name in recipes:
        del recipes[name]
        save_data(recipes, RECIPES_FILE)
        await update.message.reply_text(f"✅ Рецепт '{name}' удалён")
    else:
        await update.message.reply_text(f"Рецепт '{name}' не найден")

async def delete_all_recipes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global recipes
    recipes = {}
    save_data(recipes, RECIPES_FILE)
    await update.message.reply_text("Все рецепты удалены.")

# ---------- Расчёты ----------
async def calculate_cost(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text(
            "Укажите название десерта: /calculate название\n"
            "Пример: /calculate меренговый_рулет_белый"
        )
        return
    name = ' '.join(context.args).strip().lower()
    if name not in recipes:
        await update.message.reply_text(f"Рецепт '{name}' не найден")
        return
    data = recipes[name]
    if isinstance(data, dict) and "type" in data:
        recipe = data["ingredients"]
        portions = 1
    else:
        recipe = data["ingredients"] if isinstance(data, dict) and "ingredients" in data else data
        portions = data.get("portions", 1) if isinstance(data, dict) else 1
    total = 0.0
    missing = []
    for ing_name, qty in recipe.items():
        if ing_name in ingredients:
            total += ingredients[ing_name]["price"] * qty
        else:
            missing.append(ing_name)
    if missing:
        await update.message.reply_text(f"❌ Не хватает ингредиентов: {', '.join(missing)}.\nДобавьте их через /add_ingredient")
    else:
        base = f"💰 Себестоимость '{name}': {total:.2f} руб"
        if portions != 1:
            base += f"\n🍽 На {portions} порций: {(total/portions):.2f} руб/порция"
        await update.message.reply_text(base)

async def scale_recipe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 2:
        if len(context.args) < 2:
            await update.message.reply_text(
            "Формат: /scale название новое_количество [единица]\n"
            "Примеры:\n"
            "/scale торт 2.5 кг\n"
            "/scale рулет 3 шт"
            )
            return
    name = context.args[0].lower()
    if name not in recipes:
        await update.message.reply_text(f"Рецепт '{name}' не найден")
        return
    data = recipes[name]
    if not isinstance(data, dict) or "type" not in data:
        await update.message.reply_text("Этот рецепт нельзя масштабировать, добавьте через /add_recipe2")
        return
    try:
        if len(context.args) >= 3:
            new_qty = float(context.args[1].replace(',', '.'))
            unit = context.args[2].lower()
        else:
            match = re.match(r'^([\d.,]+)\s*([а-яa-z]+)?$', context.args[1].lower())
            if match:
                new_qty = float(match.group(1).replace(',', '.'))
                unit = match.group(2) or ''
            else:
                new_qty = float(context.args[1].replace(',', '.'))
                unit = ''
    except ValueError:
        await update.message.reply_text("Ошибка! Количество должно быть числом")
        return

    if data["type"] == "weight":
        expected_unit = "кг"
        if unit and unit not in ["кг", "килограмм", "kg"]:
            await update.message.reply_text("Для весового рецепта единица должна быть 'кг'")
            return
    else:
        expected_unit = "шт"
        if unit and unit not in ["шт", "штук", "pcs"]:
            await update.message.reply_text("Для штучного рецепта единица должна быть 'шт'")
            return

    scale_factor = new_qty / data["base_qty"]
    scaled = {ing: qty * scale_factor for ing, qty in data["ingredients"].items()}

    # --- Начало новых расчётов ---
    # Себестоимость ингредиентов
    total_ing = 0.0
    missing = []
    ing_lines = []
    for ing_name, qty in scaled.items():
        if ing_name in ingredients:
            cost = ingredients[ing_name]["price"] * qty
            total_ing += cost
            unit_i = ingredients[ing_name]["unit"]
            ing_lines.append(f"• {ing_name}: {qty:.2f} {unit_i} = {cost:.2f} руб")
        else:
            missing.append(ing_name)

    if missing:
        await update.message.reply_text(f"❌ Не хватает ингредиентов: {', '.join(missing)}")
        return

    # Дополнительные расходы (упаковка, работа)
    packaging = data.get('packaging', 0.0)
    work_hours = data.get('work_hours', 0.0)
    hourly_rate = settings.get('hourly_rate', 0.0)
    markup = data.get('markup')

    # Различаем тип рецепта: для штучных умножаем на количество, для весовых оставляем как есть
    if data["type"] == "weight":
        total_packaging = packaging  # упаковка на весь торт
        total_work = work_hours * hourly_rate if work_hours and hourly_rate else 0.0  # работа на весь торт
    else:
        total_packaging = packaging * scale_factor
        total_work = work_hours * hourly_rate * scale_factor if work_hours and hourly_rate else 0.0

    total_cost = total_ing + total_packaging + total_work

    # Формируем сообщение
    msg = f"📐 *Рецепт '{name}'* на {new_qty:.2f} {expected_unit}:\n\n"
    msg += "*Ингредиенты:*\n" + "\n".join(ing_lines)
    msg += f"\n\n💰 *Себестоимость ингредиентов:* {total_ing:.2f} руб"

    if packaging:
        msg += f"\n📦 *Упаковка:* {total_packaging:.2f} руб"
    else:
        msg += f"\n📦 *Упаковка:* не указана (0 руб)"

    if work_hours:
        if hourly_rate:
            if data["type"] == "weight":
                msg += f"\n⏱ *Работа:* {work_hours:.2f} ч × {hourly_rate:.2f} руб/ч = {total_work:.2f} руб"
            else:
                msg += f"\n⏱ *Работа:* {work_hours * scale_factor:.2f} ч × {hourly_rate:.2f} руб/ч = {total_work:.2f} руб"
        else:
            if data["type"] == "weight":
                msg += f"\n⏱ *Работа:* {work_hours:.2f} ч (ставка не задана)"
            else:
                msg += f"\n⏱ *Работа:* {work_hours * scale_factor:.2f} ч (ставка не задана)"
    else:
        msg += f"\n⏱ *Работа:* не указана (0 руб)"

    msg += f"\n🧾 *Полная себестоимость:* {total_cost:.2f} руб"

    if markup is not None:
        price = total_cost * (1 + markup / 100)
        profit = price - total_cost
        margin = (profit / total_cost) * 100 if total_cost > 0 else 0
        msg += f"\n📈 *Наценка:* {markup}%"
        msg += f"\n💵 *Цена продажи:* {price:.2f} руб"
        msg += f"\n💸 *Прибыль:* {profit:.2f} руб"
        msg += f"\n📊 *Рентабельность:* {margin:.1f}%"
    else:
        msg += f"\n❓ *Наценка не задана.* Установите через /set_markup"

    await update.message.reply_text(msg)

# ---------- Экспорт ----------
async def export_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not ingredients:
        await update.message.reply_text("Список ингредиентов пуст, нечего экспортировать")
        return
    try:
        output = StringIO()
        # Добавляем BOM для правильного отображения в Excel
        output.write('\ufeff')
        writer = csv.writer(output)
        writer.writerow(["Название", "Цена (руб)", "Единица"])
        for name, data in ingredients.items():
            writer.writerow([name, data['price'], data['unit']])
        output.seek(0)
        document = output.getvalue().encode('utf-8-sig')  # utf-8-sig автоматически добавит BOM
        await update.message.reply_document(
            document=document,
            filename="ingredients.csv",
            caption="📊 Экспорт ингредиентов"
        )
        output.close()
    except Exception as e:
        await update.message.reply_text(f"Ошибка при экспорте: {e}")

# ---------- Помощь ----------
async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("📋 Основные команды", callback_data="help_main")],
        [InlineKeyboardButton("📦 Ингредиенты и закупки", callback_data="help_ingredients")],
        [InlineKeyboardButton("🍰 Рецепты", callback_data="help_recipes")],
        [InlineKeyboardButton("💰 Продажи и аналитика", callback_data="help_sales")],
        [InlineKeyboardButton("🔄 Возвраты и списания", callback_data="help_writeoffs")],
        [InlineKeyboardButton("📅 Клиенты и заказы", callback_data="help_customers")],
        [InlineKeyboardButton("📊 Планирование и категории", callback_data="help_plans")],
        [InlineKeyboardButton("📁 Импорт/экспорт", callback_data="help_import_export")],
        [InlineKeyboardButton("⚙️ Дополнительные настройки", callback_data="help_advanced")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(
            "📖 Разделы помощи. Выберите интересующий раздел:",
            reply_markup=reply_markup
        )
    else:
        await update.message.reply_text(
            "📖 Разделы помощи. Выберите интересующий раздел:",
            reply_markup=reply_markup
        )
async def set_description(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if len(context.args) < 2:
            await update.message.reply_text(
                "Формат: /set_description рецепт описание\n"
                "Пример: /set_description меренговый_рулет_белый Взбить белки с сахаром..."
                )
            return
        recipe_name = context.args[0].lower()
        if recipe_name not in recipes:
            await update.message.reply_text(f"Рецепт '{recipe_name}' не найден")
            return
        description = ' '.join(context.args[1:])
        if isinstance(recipes[recipe_name], dict):
            recipes[recipe_name]['description'] = description
        else:
            recipes[recipe_name] = {"ingredients": recipes[recipe_name], "description": description}
        save_data(recipes, RECIPES_FILE)
        await update.message.reply_text(f"✅ Описание для рецепта '{recipe_name}' сохранено!")
    except Exception as e:
        await update.message.reply_text(f"Ошибка: {e}")
import re  # убедитесь, что этот импорт есть в начале файла

import re

async def parse_recipe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Парсит текстовый рецепт и предлагает команду для добавления"""
    if context.args:
        # Если есть аргументы, объединяем их в текст
        text = ' '.join(context.args)
        await process_recipe_text(update, context, text)
    else:
        # Если аргументов нет, просим прислать текст отдельно
        await update.message.reply_text(
            "Отправьте текст рецепта (каждый ингредиент с новой строки или через запятую):\n"
            "Например:\n"
            "мука 200 г\n"
            "сахар 150 г\n"
            "яйцо 2 шт"
        )
        context.user_data['awaiting_recipe'] = True
# Словари для временного хранения данных диалога
# Ключ: user_id
temp_recipe_data = {}

async def import_recipe_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начало диалога – отправляем приветствие и ждём текст рецепта"""
    user_id = update.effective_user.id
    temp_recipe_data[user_id] = {
        'ingredients': [],      # список кортежей (название, кол-во, ед)
        'new_ingredients': [],  # список названий новых ингредиентов
        'new_prices': {}        # словарь: название -> цена за базовую единицу
    }
    await update.message.reply_text(
        "🍰 Отправьте мне текст рецепта.\n"
        "Каждый ингредиент должен быть на отдельной строке или через запятую.\n"
        "Формат: название количество единица\n"
        "Пример:\n"
        "мука 200 г\n"
        "сахар 150 г\n"
        "яйцо 2 шт"
    )
    return WAITING_RECIPE_TEXT

async def receive_recipe_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Получаем текст рецепта, парсим ингредиенты"""
    user_id = update.effective_user.id
    text = update.message.text

    # Заменяем запятые на переносы и разбиваем на строки
    text = text.replace(',', '\n')
    lines = text.split('\n')
    units_map = {'г': 'кг', 'мл': 'л', 'кг': 'кг', 'л': 'л', 'шт': 'шт'}
    conversion = {'г': 0.001, 'мл': 0.001, 'кг': 1, 'л': 1, 'шт': 1}

    pattern = re.compile(r'^\s*([а-яА-ЯёЁa-zA-Z\s]+?)\s+(\d+[.,]?\d*)\s*(г|кг|мл|л|шт)\s*$', re.UNICODE)

    found = []
    new_ingredients = []

    for line in lines:
        line = line.strip()
        if not line:
            continue
        match = pattern.match(line)
        if match:
            ing_name_raw = match.group(1).strip().lower()
            qty_str = match.group(2).replace(',', '.')
            unit = match.group(3).lower()
            try:
                qty = float(qty_str)
            except ValueError:
                continue
            base_unit = units_map[unit]
            qty_base = qty * conversion[unit]

            # Ищем существующий ингредиент (простое сравнение)
            existing_ing = None
            for ing in ingredients:
                if ing_name_raw in ing or ing in ing_name_raw:
                    existing_ing = ing
                    break

            if existing_ing:
                ing_name = existing_ing
            else:
                ing_name = ing_name_raw
                new_ingredients.append(ing_name)

            found.append((ing_name, qty_base, base_unit))

    if not found:
        await update.message.reply_text(
            "❌ Не удалось распознать ни одного ингредиента. Попробуйте ещё раз."
        )
        return WAITING_RECIPE_TEXT

    # Сохраняем распознанные ингредиенты
    temp_recipe_data[user_id]['ingredients'] = found
    temp_recipe_data[user_id]['new_ingredients'] = new_ingredients

    # Если есть новые ингредиенты, начинаем запрашивать цены
    if new_ingredients:
        # Берём первый новый ингредиент
        next_ing = new_ingredients[0]
        unit = None
        # Находим единицу для этого ингредиента из списка
        for ing, qty, base_unit in found:
            if ing == next_ing:
                unit = base_unit
                break
        if not unit:
            unit = 'кг'  # запасной вариант
        await update.message.reply_text(
            f"🆕 Найден новый ингредиент: *{next_ing}*.\n"
            f"Введите цену за 1 {unit} (например, 150):"
        )
        return WAITING_INGREDIENT_PRICE
    else:
        # Если все ингредиенты известны, переходим к запросу названия рецепта
        await update.message.reply_text("Введите название рецепта (например, «Красный бархат»):")
        return WAITING_RECIPE_NAME

async def receive_ingredient_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Получаем цену для очередного нового ингредиента"""
    user_id = update.effective_user.id
    text = update.message.text.strip()

    try:
        price = float(text.replace(',', '.'))
    except ValueError:
        await update.message.reply_text("❌ Ошибка! Введите число (например, 150).")
        return WAITING_INGREDIENT_PRICE

    # Определяем, для какого ингредиента мы ждём цену
    new_ings = temp_recipe_data[user_id]['new_ingredients']
    if not new_ings:
        # такого не должно быть, но на всякий случай
        await update.message.reply_text("Введите название рецепта:")
        return WAITING_RECIPE_NAME

    current_ing = new_ings[0]
    temp_recipe_data[user_id]['new_prices'][current_ing] = price

    # Удаляем обработанный ингредиент из списка новых
    new_ings.pop(0)

    if new_ings:
        # Есть ещё новые ингредиенты
        next_ing = new_ings[0]
        # Находим его единицу
        unit = None
        for ing, qty, base_unit in temp_recipe_data[user_id]['ingredients']:
            if ing == next_ing:
                unit = base_unit
                break
        if not unit:
            unit = 'кг'
        await update.message.reply_text(
            f"🆕 Следующий новый ингредиент: *{next_ing}*.\n"
            f"Введите цену за 1 {unit}:"
        )
        return WAITING_INGREDIENT_PRICE
    else:
        # Все цены введены, переходим к названию рецепта
        await update.message.reply_text("Введите название рецепта (например, «Красный бархат»):")
        return WAITING_RECIPE_NAME

async def receive_recipe_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Получаем название рецепта, запрашиваем тип"""
    user_id = update.effective_user.id
    recipe_name = update.message.text.strip().lower()
    temp_recipe_data[user_id]['recipe_name'] = recipe_name

    await update.message.reply_text(
        "Введите тип рецепта и базовое количество.\n"
        "Примеры:\n"
        "- для весового: вес 1 кг\n"
        "- для штучного: штук 1 шт\n"
        "Вы можете указать другое базовое количество, например: вес 0.5 кг"
    )
    return WAITING_RECIPE_TYPE

async def receive_recipe_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Получаем тип рецепта и базовое количество, сохраняем всё в базу"""
    user_id = update.effective_user.id
    text = update.message.text.strip().lower()

    # Парсим тип и количество
    parts = text.split()
    if len(parts) < 2:
        await update.message.reply_text(
            "❌ Неверный формат. Введите, например: вес 1 кг или штук 1 шт"
        )
        return WAITING_RECIPE_TYPE

    type_str = parts[0]
    try:
        base_qty = float(parts[1].replace(',', '.'))
    except ValueError:
        await update.message.reply_text("❌ Ошибка! Количество должно быть числом.")
        return WAITING_RECIPE_TYPE

    if type_str in ['вес', 'weight']:
        recipe_type = 'weight'
        if len(parts) >= 3 and parts[2] in ['кг', 'килограмм', 'kg']:
            base_unit = 'кг'
        else:
            base_unit = 'кг'  # по умолчанию
    elif type_str in ['штук', 'pcs', 'шт']:
        recipe_type = 'pcs'
        if len(parts) >= 3 and parts[2] in ['шт', 'штук', 'pcs']:
            base_unit = 'шт'
        else:
            base_unit = 'шт'
    else:
        await update.message.reply_text("❌ Тип должен быть «вес» или «штук».")
        return WAITING_RECIPE_TYPE

    # Добавляем новые ингредиенты в базу с указанными ценами
    new_prices = temp_recipe_data[user_id]['new_prices']
    for ing_name, price in new_prices.items():
        # Определяем единицу для этого ингредиента (из ранее распознанного списка)
        unit = 'кг'  # по умолчанию
        for ing, qty, base_unit in temp_recipe_data[user_id]['ingredients']:
            if ing == ing_name:
                unit = base_unit
                break
        ingredients[ing_name] = {"price": price, "unit": unit, "stock": 0.0}
        save_data(ingredients, INGREDIENTS_FILE)

    # Формируем словарь ингредиентов для рецепта
    recipe_ingredients = {}
    for ing, qty, unit in temp_recipe_data[user_id]['ingredients']:
        recipe_ingredients[ing] = qty

    # Сохраняем рецепт
    recipe_data = {
        "type": recipe_type,
        "base_qty": base_qty,
        "ingredients": recipe_ingredients
    }
    recipe_name = temp_recipe_data[user_id]['recipe_name']
    recipes[recipe_name] = recipe_data
    save_data(recipes, RECIPES_FILE)

    # Очищаем временные данные
    del temp_recipe_data[user_id]

    await update.message.reply_text(
        f"✅ Рецепт «{recipe_name}» успешно добавлен!\n"
        f"Новые ингредиенты также добавлены в базу."
    )
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отмена диалога"""
    user_id = update.effective_user.id
    if user_id in temp_recipe_data:
        del temp_recipe_data[user_id]
    await update.message.reply_text("❌ Диалог отменён.")
    return ConversationHandler.END
async def process_recipe_text(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
    """Основная логика парсинга"""
    # Заменяем запятые на переносы строк и разбиваем
    print(f"process_recipe_text called with text: {text}")
    lines = text.strip().split('\n')
    print(f"lines: {lines}")
    units_map = {'г': 'кг', 'мл': 'л', 'кг': 'кг', 'л': 'л', 'шт': 'шт'}
    conversion = {'г': 0.001, 'мл': 0.001, 'кг': 1, 'л': 1, 'шт': 1}

    # Регулярное выражение: название (может состоять из нескольких слов), затем число, затем единица
    pattern = re.compile(r'^\s*([а-яА-ЯёЁa-zA-Z\s]+?)\s+(\d+[.,]?\d*)\s*(г|кг|мл|л|шт)\s*$', re.UNICODE)

    for line in lines:
        line = line.strip()
        if not line:
            continue
        match = pattern.match(line)
        if match:
            ing_name_raw = match.group(1).strip().lower()
            qty_str = match.group(2).replace(',', '.')
            unit = match.group(3).lower()
            try:
                qty = float(qty_str)
            except ValueError:
                continue
            base_unit = units_map[unit]
            qty_base = qty * conversion[unit]

            # Ищем существующий ингредиент (простое сравнение)
            found_ing = None
            for existing in ingredients:
                if ing_name_raw in existing or existing in ing_name_raw:
                    found_ing = existing
                    break
            if not found_ing:
                found_ing = ing_name_raw  # используем как есть

            found.append((found_ing, qty_base, base_unit))
        else:
            # Если строка не распознана, игнорируем
            pass

    if not found:
        await update.message.reply_text(
            "❌ Не удалось распознать ингредиенты.\n"
            "Убедитесь, что каждый ингредиент указан в формате:\n"
            "название количество единица\n"
            "Например: мука 200 г\n"
            "Допустимые единицы: г, кг, мл, л, шт"
        )
        return

    # Формируем команду /add_recipe2
    ing_parts = []
    for ing, qty, unit in found:
        # Красивое форматирование числа (убираем лишние нули)
        qty_str = f"{qty:.3f}".rstrip('0').rstrip('.') if '.' in f"{qty:.3f}" else f"{qty:.3f}"
        ing_parts.append(f"{ing} {qty_str}")

    cmd = f"/add_recipe2 новый_рецепт штук 1: " + ", ".join(ing_parts)

    msg = "✅ *Распознанные ингредиенты:*\n"
    for (ing, qty, unit) in found:
        msg += f"• {ing}: {qty:.3f} {unit}\n"
    msg += f"\n📝 *Команда для добавления*\n(проверьте и скорректируйте название рецепта):\n`{cmd}`"

    await update.message.reply_text(msg)
async def plan_recipe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 3:
        await update.message.reply_text(
        "Формат: /plan рецепт количество дата\n"
        "Пример: /plan меренговый_рулет_белый 5 2025-03-15"
        )
        return
    *name_parts, qty_str, date_str = context.args
    name = ' '.join(name_parts).lower()
    try:
        qty = float(qty_str.replace(',', '.'))
        date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        await update.message.reply_text("Ошибка! Количество должно быть числом, дата в формате ГГГГ-ММ-ДД")
        return
    if name not in recipes:
        await update.message.reply_text(f"Рецепт '{name}' не найден")
        return
    plan = {
        "date": date.isoformat(),
        "recipe": name,
        "quantity": qty
    }
    plans.append(plan)
    save_plans()
    await update.message.reply_text(f"✅ Запланировано {qty} шт '{name}' на {date_str}")
async def plan_recipe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Запланировать приготовление: /plan рецепт количество ГГГГ-ММ-ДД"""
    if len(context.args) < 3:
        await update.message.reply_text(
            "Формат: /plan рецепт количество дата\n"
            "Пример: /plan меренговый_рулет_белый 5 2025-03-15"
        )
        return
    # Собираем название рецепта (может быть из нескольких слов)
    *name_parts, qty_str, date_str = context.args
    name = ' '.join(name_parts).lower()
    try:
        qty = float(qty_str.replace(',', '.'))
        date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        await update.message.reply_text("Ошибка! Количество должно быть числом, дата в формате ГГГГ-ММ-ДД")
        return
    if name not in recipes:
        await update.message.reply_text(f"Рецепт '{name}' не найден")
        return
    plan = {
        "date": date.isoformat(),
        "recipe": name,
        "quantity": qty
    }
    plans.append(plan)
    save_plans()
    await update.message.reply_text(f"✅ Запланировано {qty} шт '{name}' на {date_str}")
# ========== Команда /shopping (список покупок) ==========
async def shopping(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает список покупок на основе заказов (без Markdown)"""
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)

    # Собираем заказы на сегодня и завтра
    upcoming_orders = []
    for order in orders:
        if 'date' in order:
            order_date = datetime.strptime(order['date'], "%Y-%m-%d").date()
        elif 'due_date' in order:
            order_date = datetime.strptime(order['due_date'], "%Y-%m-%d").date()
        else:
            continue
        if order_date == today or order_date == tomorrow:
            upcoming_orders.append(order)

    if not upcoming_orders:
        await update.message.reply_text("✅ Нет заказов на сегодня и завтра.")
        return

    needs = {}
    for order in upcoming_orders:
        recipe_name = order['recipe']
        qty = order.get('quantity', 1)
        if recipe_name not in recipes:
            continue
        data = recipes[recipe_name]
        if isinstance(data, dict) and "ingredients" in data:
            ing_dict = data["ingredients"]
        else:
            ing_dict = data
        for ing_name, ing_qty in ing_dict.items():
            need = ing_qty * qty
            needs[ing_name] = needs.get(ing_name, 0.0) + need

    to_buy = {}
    for ing_name, need in needs.items():
        stock = ingredients.get(ing_name, {}).get('stock', 0.0)
        deficit = need - stock
        if deficit > 0:
            to_buy[ing_name] = deficit

    if not to_buy:
        await update.message.reply_text("✅ У вас достаточно ингредиентов для выполнения заказов.")
        return

    msg = "🛒 Список покупок\n\n"
    for ing_name, need in sorted(to_buy.items()):
        unit = ingredients.get(ing_name, {}).get('unit', '')
        msg += f"• {ing_name}: {need:.2f} {unit}\n"
    await update.message.reply_text(msg)
async def process_recipe_text(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
    """Основная логика парсинга"""
    # Простейший парсер: ищем строки вида "ингредиент число единица"
    lines = text.strip().split('\n')
    found = []
    units_map = {'г': 'кг', 'мл': 'л', 'кг': 'кг', 'л': 'л', 'шт': 'шт'}
    conversion = {'г': 0.001, 'мл': 0.001, 'кг': 1, 'л': 1, 'шт': 1}

    # Регулярное выражение: слово(а) + пробел + число (с точкой/запятой) + пробел + единица
    pattern = re.compile(r'^\s*([а-яА-ЯёЁa-zA-Z\s]+?)\s+(\d+[.,]?\d*)\s*(г|кг|мл|л|шт)\s*$', re.UNICODE)

    for line in lines:
        line = line.strip()
        if not line:
            continue
        match = pattern.match(line)
        if match:
            ing_name_raw = match.group(1).strip().lower()
            qty_str = match.group(2).replace(',', '.')
            unit = match.group(3).lower()
            try:
                qty = float(qty_str)
            except ValueError:
                continue
            # Переводим в базовую единицу
            base_unit = units_map[unit]
            qty_base = qty * conversion[unit]

            # Ищем похожий ингредиент в базе (простое сравнение по нижнему регистру)
            found_ing = None
            for existing in ingredients:
                if ing_name_raw in existing or existing in ing_name_raw:
                    found_ing = existing
                    break
            if not found_ing:
                # Если не нашли, предлагаем создать новый
                found_ing = ing_name_raw
                # Добавим в список для дальнейшего использования (но пока не сохраняем)
                # Можно было бы предложить создать ингредиент отдельно, но упростим: оставляем как есть
                pass

            found.append((found_ing, qty_base, base_unit))
        else:
            # Если строка не распознана, просто игнорируем (можно уведомить)
            pass

    if not found:
        await update.message.reply_text("Не удалось распознать ингредиенты. Проверьте формат:\nназвание количество единица\nНапример: мука 200 г")
        return

    # Формируем команду /add_recipe2
    ing_parts = []
    for ing, qty, unit in found:
        ing_parts.append(f"{ing} {qty:.3f}".rstrip('0').rstrip('.') if '.' in f"{qty:.3f}" else f"{qty:.3f}")
    cmd = f"/add_recipe2 новый_рецепт штук 1: " + ", ".join(ing_parts)

    # Отправляем результат пользователю
    msg = "✅ Распознанные ингредиенты:\n"
    for (ing, qty, unit) in found:
        msg += f"• {ing}: {qty:.3f} {unit}\n"
    msg += f"\nКоманда для добавления (проверьте и скорректируйте название рецепта):\n`{cmd}`"
    await update.message.reply_text(msg)
# ---------- Продвинутый показ рецепта ----------
async def show_recipe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text(
           "Укажите название рецепта: /show_recipe название\n"
            "Пример: /show_recipe меренговый_рулет_белый"
        )
        return
    name = ' '.join(context.args).strip().lower()
    if name not in recipes:
        await update.message.reply_text(f"Рецепт '{name}' не найден")
        return
    data = recipes[name]
    # Получаем ингредиенты
    if isinstance(data, dict) and "ingredients" in data:
        ing_dict = data["ingredients"]
        base_info = ""
        if "type" in data:
            base_info = f"Тип: {data['type']}, база: {data['base_qty']} {'кг' if data['type']=='weight' else 'шт'}\n"
    else:
        ing_dict = data
        base_info = "Старый формат рецепта (без масштабирования)\n"
    # Рассчитываем себестоимость ингредиентов
    total_ing = 0.0
    missing = []
    for ing_name, qty in ing_dict.items():
        if ing_name in ingredients:
            total_ing += ingredients[ing_name]["price"] * qty
        else:
            missing.append(ing_name)
    # Формируем основное сообщение (без описания)
    msg = f"🍰 {name}\n"
    msg += base_info
    msg += "\nИнгредиенты:\n"
    for ing_name, qty in ing_dict.items():
        unit = ingredients.get(ing_name, {}).get('unit', '')
        msg += f"• {ing_name}: {qty} {unit}\n"
    if missing:
        msg += f"\n⚠️ Отсутствуют в базе: {', '.join(missing)}\n"
        await update.message.reply_text(msg)
        return
    # Дополнительные расходы
    packaging = data.get('packaging') if isinstance(data, dict) else None
    work_hours = data.get('work_hours') if isinstance(data, dict) else None
    markup = data.get('markup') if isinstance(data, dict) else None
    hourly_rate = settings.get('hourly_rate', 0.0)
    work_cost = work_hours * hourly_rate if work_hours is not None and hourly_rate > 0 else None
    total_cost = total_ing
    if packaging is not None:
        total_cost += packaging
    if work_cost is not None:
        total_cost += work_cost
    msg += f"\n💰 Себестоимость ингредиентов: {total_ing:.2f} руб"
    if packaging is not None:
        msg += f"\n📦 Упаковка: {packaging:.2f} руб"
    else:
        msg += f"\n📦 Упаковка: не указана (0 руб)"
    if work_hours is not None:
        if hourly_rate > 0:
            msg += f"\n⏱ Работа: {work_hours:.2f} ч × {hourly_rate:.2f} руб/ч = {work_cost:.2f} руб"
        else:
            msg += f"\n⏱ Работа: {work_hours:.2f} ч (ставка не задана)"
    else:
        msg += f"\n⏱ Работа: не указана (0 руб)"
    msg += f"\n🧾 Полная себестоимость: {total_cost:.2f} руб"
    if markup is not None:
        price = total_cost * (1 + markup/100)
        profit = price - total_cost
        margin = (profit / total_cost) * 100 if total_cost > 0 else 0
        msg += f"\n📈 Наценка: {markup}%"
        msg += f"\n💵 Цена продажи: {price:.2f} руб"
        msg += f"\n💸 Прибыль: {profit:.2f} руб"
        msg += f"\n📊 Рентабельность: {margin:.1f}%"
    else:
        msg += f"\n❓ Наценка не задана. Установите через /set_markup"

    # Функция для отправки длинных сообщений частями
    async def send_long_message(text):
        if len(text) <= 4096:
            await update.message.reply_text(text)
        else:
            parts = [text[i:i+4096] for i in range(0, len(text), 4096)]
            for part in parts:
                await update.message.reply_text(part)

    # Отправляем основное сообщение
    await send_long_message(msg)

    # Если есть описание, отправляем его отдельно
    if isinstance(data, dict) and "description" in data:
        desc_msg = f"🍳 Приготовление:\n{data['description']}"
        await send_long_message(desc_msg)

# ---------- Установка почасовой ставки ----------
async def set_hourly_rate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Укажите ставку: /set_hourly_rate 350")
        return
    try:
        rate = float(context.args[0].replace(',', '.'))
        settings['hourly_rate'] = rate
        save_settings()
        await update.message.reply_text(f"✅ Почасовая ставка установлена: {rate} руб/час")
    except ValueError:
        await update.message.reply_text("Ошибка! Ставка должна быть числом")

# ---------- Установка стоимости упаковки ----------
async def set_packaging(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 2:
        await update.message.reply_text(
        "Формат: /set_packaging название цена\n"
        "Пример: /set_packaging меренговый_рулет_белый 76"
        )
        return
    *name_parts, price_str = context.args
    name = ' '.join(name_parts).lower()
    if name not in recipes:
        await update.message.reply_text(f"Рецепт '{name}' не найден")
        return
    try:
        price = float(price_str.replace(',', '.'))
        if isinstance(recipes[name], dict):
            recipes[name]['packaging'] = price
        else:
            recipes[name] = {"ingredients": recipes[name], "packaging": price}
        save_data(recipes, RECIPES_FILE)
        await update.message.reply_text(f"✅ Для рецепта '{name}' установлена упаковка: {price} руб")
    except ValueError:
        await update.message.reply_text("Ошибка! Цена должна быть числом")

# ---------- Установка времени работы ----------
async def set_work_hours(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 2:
        await update.message.reply_text(
        "Формат: /set_work_hours название часы\n"
        "Пример: /set_work_hours меренговый_рулет_белый 0.75"
        )
        return
    *name_parts, hours_str = context.args
    name = ' '.join(name_parts).lower()
    if name not in recipes:
        await update.message.reply_text(f"Рецепт '{name}' не найден")
        return
    try:
        hours = float(hours_str.replace(',', '.'))
        if isinstance(recipes[name], dict):
            recipes[name]['work_hours'] = hours
        else:
            recipes[name] = {"ingredients": recipes[name], "work_hours": hours}
        save_data(recipes, RECIPES_FILE)
        await update.message.reply_text(f"✅ Для рецепта '{name}' установлено время работы: {hours} ч")
    except ValueError:
        await update.message.reply_text("Ошибка! Часы должны быть числом")

# ---------- Установка наценки ----------
async def set_markup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 2:
        await update.message.reply_text(
        "Формат: /set_markup название процент\n"
        "Пример: /set_markup меренговый_рулет_белый 40"
        )
        return
    *name_parts, markup_str = context.args
    name = ' '.join(name_parts).lower()
    if name not in recipes:
        await update.message.reply_text(f"Рецепт '{name}' не найден")
        return
    try:
        markup = float(markup_str.replace(',', '.'))
        if isinstance(recipes[name], dict):
            recipes[name]['markup'] = markup
        else:
            recipes[name] = {"ingredients": recipes[name], "markup": markup}
        save_data(recipes, RECIPES_FILE)
        await update.message.reply_text(f"✅ Для рецепта '{name}' установлена наценка: {markup}%")
    except ValueError:
        await update.message.reply_text("Ошибка! Процент должен быть числом")
async def add_stock(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 2:
        await update.message.reply_text(
            "Формат: /add_stock название количество\n"
            "Пример: /add_stock мука 10"
        )
        return
    *name_parts, qty_str = context.args
    name = ' '.join(name_parts).lower()
    if name not in ingredients:
        await update.message.reply_text(f"Ингредиент '{name}' не найден")
        return
    try:
        qty = float(qty_str.replace(',', '.'))
        # Если поле stock отсутствует, создаём его
        if 'stock' not in ingredients[name]:
            ingredients[name]['stock'] = 0.0
        ingredients[name]['stock'] += qty
        save_data(ingredients, INGREDIENTS_FILE)
        await update.message.reply_text(
            f"✅ Добавлено {qty} {ingredients[name]['unit']} к '{name}'. "
            f"Текущий остаток: {ingredients[name]['stock']:.2f} {ingredients[name]['unit']}"
        )
    except ValueError:
        await update.message.reply_text("Ошибка! Количество должно быть числом")
# ---------- Список рецептов с ценами ----------
async def price_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not recipes:
        await update.message.reply_text("Список рецептов пуст")
        return
    msg = "📋 *Список рецептов с ценами:*\n\n"
    hourly_rate = settings.get('hourly_rate', 0.0)
    for name, data in recipes.items():
        # Себестоимость ингредиентов
        if isinstance(data, dict) and "ingredients" in data:
            ing_dict = data["ingredients"]
        else:
            ing_dict = data
        total_ing = 0.0
        missing = False
        for ing_name, qty in ing_dict.items():
            if ing_name in ingredients:
                total_ing += ingredients[ing_name]["price"] * qty
            else:
                missing = True
                break
        if missing:
            continue  # Пропускаем рецепты с отсутствующими ингредиентами
        packaging = data.get('packaging') if isinstance(data, dict) else None
        work_hours = data.get('work_hours') if isinstance(data, dict) else None
        markup = data.get('markup') if isinstance(data, dict) else None
        total_cost = total_ing
        if packaging:
            total_cost += packaging
        if work_hours and hourly_rate > 0:
            total_cost += work_hours * hourly_rate
        price = total_cost * (1 + markup/100) if markup else None
        msg += f"• *{name}*"
        if price:
            msg += f": {price:.2f} руб (себ. {total_cost:.2f} руб)"
        else:
            msg += f": себест. {total_cost:.2f} руб (наценка не задана)"
        msg += "\n"
    await update.message.reply_text(msg)
# ---------- Обработчик кнопок ----------
async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    # Игнорируем callback'и, которые обрабатываются другими обработчиками (например, /price)
    if data.startswith('price_'):
        return

    # Старые обработчики (из вашего кода)
    if data == 'add_ing':
        await query.edit_message_text(
            "Чтобы добавить ингредиент, отправьте команду:\n"
            "/add_ingredient название цена единица\n"
            "Или с ценой за упаковку: /add_ingredient название цена_упаковки вес_упаковки единица\n\n"
            "Допустимые единицы: кг, г, шт, л, мл"
        )
    elif data == 'list_ing':
        await show_ingredients(update, context)
    elif data == 'add_rcp':
        await query.edit_message_text(
            "Чтобы добавить рецепт, отправьте:\n"
            "/add_recipe Название: порции; ингредиенты (старый формат)\n"
            "или /add_recipe2 название тип базовое_количество: ингредиенты (для масштабирования)\n\n"
            "Примеры:\n"
            "/add_recipe Омлет: 2; яйца 3, молоко 0.1\n"
            "/add_recipe2 торт вес 1: мука 0.5, сахар 0.2, яйца 3"
        )
    elif data == 'calc':
        await query.edit_message_text(
            "Введите название десерта для расчёта:\n"
            "/calculate название\n"
            "Например: /calculate омлет"
        )
    elif data == 'list_rcp':
        await list_recipes(update, context)
    elif data == 'scale':
        await query.edit_message_text(
            "Чтобы пересчитать рецепт на нужный вес/количество:\n"
            "/scale название новое_количество [единица]\n\n"
            "Примеры:\n"
            "/scale торт 2.5 кг\n"
            "/scale печенье 30 шт"
        )
    elif data == 'help':
        await help_command(update, context)
    # Добавьте сюда другие обработчики, если они были в вашем старом button_handler
    else:
        # Если ни одно условие не подошло, можно ничего не делать или отправить сообщение
        # В старом коде у вас мог быть такой else, но теперь он не должен срабатывать для price_*
        print(f"Неизвестный callback: {data}")
async def export_full(update: Update, context: ContextTypes.DEFAULT_TYPE):
    import zipfile, io
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        def add_json(data, filename):
            json_str = json.dumps(data, ensure_ascii=False, indent=2)
            zip_file.writestr(filename, json_str.encode('utf-8'))
        add_json(ingredients, "ingredients.json")
        add_json(recipes, "recipes.json")
        add_json(settings, "settings.json")
        add_json(sales, "sales.json")
        add_json(customers, "customers.json")
        add_json(orders, "orders.json")
        add_json(plans, "plans.json")
    zip_buffer.seek(0)
    await update.message.reply_document(
        document=zip_buffer,
        filename="backup.zip",
        caption="📦 Полный бэкап данных"
    )
async def list_categories(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показать все категории, которые есть в рецептах"""
    cats = set()
    for data in recipes.values():
        if isinstance(data, dict) and 'category' in data:
            cats.add(data['category'])
    if not cats:
        await update.message.reply_text("Нет категорий.")
        return
    msg = "📂 *Категории:*\n" + "\n".join(f"• {c}" for c in sorted(cats))
    await update.message.reply_text(msg)
async def export_xlsx(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not ingredients:
        await update.message.reply_text("Список ингредиентов пуст, нечего экспортировать")
        return
    try:
        # Создаём книгу и лист
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ингредиенты"
        
        # Заголовки
        headers = ["Название", "Цена (руб)", "Единица"]
        ws.append(headers)
        # Делаем заголовки жирными
        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
        
        # Данные
        for name, data in ingredients.items():
            ws.append([name, data['price'], data['unit']])
        
        # Автоподбор ширины колонок
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Сохраняем в буфер
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        await update.message.reply_document(
            document=output,
            filename="ingredients.xlsx",
            caption="📊 Экспорт ингредиентов в формате Excel"
        )
        output.close()
    except Exception as e:
        await update.message.reply_text(f"Ошибка при экспорте: {e}")
async def report_xlsx(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Выгрузить отчёт (план vs факт) за период в Excel: /report_xlsx [день|неделя|месяц|год]"""
    period = context.args[0].lower() if context.args else "месяц"
    now = datetime.now()
    if period == "день":
        start = now - timedelta(days=1)
    elif period == "неделя":
        start = now - timedelta(weeks=1)
    elif period == "месяц":
        start = now - timedelta(days=30)
    elif period == "год":
        start = now - timedelta(days=365)
    else:
        await update.message.reply_text("Период может быть: день, неделя, месяц, год")
        return

    # Собираем планы за период
    plans_in_period = {}
    for plan in plans:
        plan_date = datetime.fromisoformat(plan['date'])
        if plan_date >= start:
            recipe = plan['recipe']
            qty = plan['quantity']
            plans_in_period[recipe] = plans_in_period.get(recipe, 0.0) + qty

    # Собираем продажи за период
    sales_in_period = {}
    profit_by_recipe = {}
    total_profit = 0.0
    for sale in sales:
        sale_date = datetime.fromisoformat(sale['date'])
        if sale_date >= start:
            recipe = sale['recipe']
            qty = sale.get('quantity', 1)
            profit = sale.get('profit', 0.0) or 0.0
            sales_in_period[recipe] = sales_in_period.get(recipe, 0.0) + qty
            profit_by_recipe[recipe] = profit_by_recipe.get(recipe, 0.0) + profit
            total_profit += profit

    if not plans_in_period and not sales_in_period:
        await update.message.reply_text(f"Нет данных за {period}.")
        return

    # Создаём Excel-файл
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Отчёт {period}"

        # Заголовки
        headers = ["Рецепт", "План", "Факт", "Разница", "Прибыль"]
        ws.append(headers)
        for col in range(1, 6):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Данные
        all_recipes = set(plans_in_period.keys()) | set(sales_in_period.keys())
        row = 2
        for recipe in sorted(all_recipes):
            plan = plans_in_period.get(recipe, 0)
            fact = sales_in_period.get(recipe, 0)
            diff = fact - plan
            profit = profit_by_recipe.get(recipe, 0)
            ws.append([recipe, plan, fact, diff, profit])
            # Выравнивание чисел
            for col in range(2, 6):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
            row += 1

        # Итоговая строка
        ws.append(["Итого", "", "", "", total_profit])
        ws.cell(row=row, column=5).font = Font(bold=True)
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')

        # Автоширина колонок
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # Сохраняем в буфер
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"report_{period}_{now.strftime('%Y%m%d')}.xlsx"
        await update.message.reply_document(
            document=output,
            filename=filename,
            caption=f"📊 Отчёт за {period}"
        )
        output.close()

    except Exception as e:
        await update.message.reply_text(f"Ошибка при создании отчёта: {e}")
async def scheduled_remind(context: ContextTypes.DEFAULT_TYPE):
    """Отправляет напоминание о заказах на завтра (вызывается по расписанию)"""
    tomorrow = datetime.now().date() + timedelta(days=1)
    admin_chat_id = settings.get('admin_chat_id')
    if not admin_chat_id:
        print("[DEBUG] admin_chat_id не задан")
        return

    msg = f"🔔 Напоминание о заказах на ЗАВТРА ({tomorrow}):\n"
    found = False
    for order in orders:
        order_date = datetime.fromisoformat(order['due_date']).date()
        if order_date == tomorrow:
            msg += f"• {order['customer']} – {order['recipe']} ({order['quantity']} шт)\n"
            found = True

    if not found:
        msg += "Нет заказов на завтра."

    await context.bot.send_message(chat_id=admin_chat_id, text=msg)
    print(f"[DEBUG] Уведомление отправлено в чат {admin_chat_id}")
def record_price_history(ingredient, old_price, new_price, unit):
    """Сохраняет запись об изменении цены с отладкой"""
    print(f"🔥 record_price_history вызвана: ингредиент={ingredient}, old={old_price}, new={new_price}, unit={unit}")
    # Защита от Ellipsis
    if old_price is ...:
        old_price = None
        print("⚠️ old_price был Ellipsis")
    if new_price is ...:
        new_price = None
        print("⚠️ new_price был Ellipsis")
    try:
        record = {
            "date": datetime.now().isoformat(),
            "ingredient": ingredient,
            "old_price": old_price,
            "new_price": new_price,
            "unit": unit
        }
        price_history.append(record)
        save_price_history()
        print("✅ История сохранена")
    except Exception as e:
        print(f"❌ Ошибка в record_price_history: {e}")
        raise  # пробрасываем дальше, чтобы увидеть в основном коде
async def purchase(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Добавить закупку ингредиента с указанием срока годности"""
    args = context.args
    if len(args) < 4:
        await update.message.reply_text(
            "Формат: /purchase ингредиент количество цена ГГГГ-ММ-ДД [поставщик]\n"
            "Пример: /purchase мука 10 500 2026-06-01 Мельница\n"
            "Примечание: название ингредиента может состоять из нескольких слов."
        )
        return

    # Последние 4 аргумента: количество, цена, дата, и возможно поставщик?
    # Мы знаем, что после названия идут: количество, цена, дата, и опционально поставщик.
    # Поэтому отделим последние 3 аргумента как обязательные (количество, цена, дата),
    # а всё, что до них — название.
    # Если аргументов ровно 4, то последний — дата, а названия — первые.
    # Если аргументов 5, то последний — поставщик, предпоследний — дата, и т.д.
    # Но проще: количество, цена, дата — это последние 3 аргумента.
    # Всё, что до них — название.
    # Потом проверим, есть ли ещё один аргумент после даты? Нет, дата последняя, если нет поставщика.
    # Если есть поставщик, то аргументов должно быть на 1 больше, и поставщик будет последним.
    # То есть: название (N слов) + количество, цена, дата = N+3 аргументов.
    # Если есть поставщик, то +1 = N+4.

    # Определим базовые аргументы: количество, цена, дата — последние 3
    qty_str = args[-3]
    price_str = args[-2]
    date_str = args[-1]
    name_parts = args[:-3]  # всё до последних трёх

    # Проверяем, есть ли лишний аргумент? Если длина args больше, чем name_parts + 3, значит есть поставщик.
    # Но name_parts уже включает всё до последних трёх. Если есть поставщик, то name_parts будет содержать его?
    # Нет, потому что поставщик — это последний аргумент, если он есть. А мы отделили последние три, значит поставщик мог остаться в name_parts? 
    # Давайте подойдём иначе: выделим последние 4 аргумента, если аргументов >=5? Нет, проще: предположим, что поставщик может быть только один и он идёт последним.
    # Если аргументов ровно 4, то поставщика нет.
    # Если аргументов 5, то поставщик есть и он последний.
    # Если аргументов больше 5, то название многословное и поставщик всё равно последний.

    # Найдём, где могут быть количество, цена, дата. Они всегда последние три.
    # Поставщик, если есть, должен стоять перед датой? Нет, по формату он последний.
    # Значит, возможны варианты:
    # - args = [name..., qty, price, date]  (без поставщика)
    # - args = [name..., qty, price, date, supplier] (с поставщиком)
    # В обоих случаях последние три аргумента — qty, price, date.
    # А supplier, если есть, будет args[-4]? Нет, если supplier есть, то последние четыре: qty, price, date, supplier? Неправильно, supplier последний, значит перед датой? Перепутали.
    # Правильно: порядок: название, количество, цена, дата, [поставщик]. То есть дата предпоследняя, поставщик последний.
    # Тогда последние два аргумента — дата и поставщик? Нет, последние два: дата и поставщик, если поставщик есть. А количество и цена перед ними.
    # Значит, общая картина: название (слова) + qty + price + date + (supplier).
    # Тогда количество аргументов = len(name_parts) + 3 + (1 если supplier есть).
    # Но мы не знаем, supplier есть или нет.

    # Более простой способ: попытаемся распарсить, предполагая, что последний аргумент может быть датой или поставщиком.
    # Проверим, является ли последний аргумент датой (формат ГГГГ-ММ-ДД). Если да, то поставщика нет, и последние три — qty, price, date.
    # Если последний не дата, значит это поставщик, тогда дата — предпоследний.

    # Реализуем это:
    try:
        # Проверяем последний аргумент на дату
        datetime.strptime(args[-1], "%Y-%m-%d")
        # Если успешно, значит последний — дата, поставщика нет
        date_str = args[-1]
        price_str = args[-2]
        qty_str = args[-3]
        name_parts = args[:-3]
        supplier = ""
    except ValueError:
        # Последний не дата, значит это поставщик, а дата предпоследний
        date_str = args[-2]
        price_str = args[-3]
        qty_str = args[-4]
        name_parts = args[:-4]
        supplier = args[-1]

    name = ' '.join(name_parts).lower()
    if not name:
        await update.message.reply_text("Не указано название ингредиента.")
        return

    try:
        qty = float(qty_str.replace(',', '.'))
        price = float(price_str.replace(',', '.'))
        expiry = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        await update.message.reply_text("Ошибка! Проверьте количество, цену и формат даты (ГГГГ-ММ-ДД).")
        return

    if name not in ingredients:
        await update.message.reply_text(f"Ингредиент '{name}' не найден. Сначала добавьте через /add_ingredient")
        return

    # Создаём новую партию
    batch = {
        "id": f"batch_{uuid.uuid4().hex[:8]}",
        "ingredient": name,
        "initial_quantity": qty,
        "current_quantity": qty,
        "unit": ingredients[name]["unit"],
        "purchase_date": datetime.now().date().isoformat(),
        "expiry_date": expiry.isoformat(),
        "cost": price,
        "supplier": supplier,
        "is_active": True
    }
    batches.append(batch)
    save_batches()

    # Обновляем общий остаток в ingredients
    if 'stock' not in ingredients[name]:
        ingredients[name]['stock'] = 0.0
    ingredients[name]['stock'] += qty
    save_data(ingredients, INGREDIENTS_FILE)

    await update.message.reply_text(
        f"✅ Закупка '{name}' ({qty} {ingredients[name]['unit']}) за {price} руб зарегистрирована.\n"
        f"Срок годности до {expiry}. Поставщик: {supplier or 'не указан'}."
    )
async def expiry_check(context: ContextTypes.DEFAULT_TYPE):
    """Проверяет сроки годности и отправляет уведомление"""
    admin_chat_id = settings.get('admin_chat_id')
    if not admin_chat_id:
        return
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)
    yesterday = today - timedelta(days=1)

    expired = []
    expiring_today = []
    expiring_tomorrow = []

    for batch in batches:
        if not batch.get('is_active', True):
            continue
        if batch.get('opened_date'):
            exp_date = datetime.fromisoformat(batch['expiry_after_open']).date()
        else:
            exp_date = datetime.fromisoformat(batch['expiry_date']).date()
        if exp_date < today:
            expired.append(batch)
        elif exp_date == today:
            expiring_today.append(batch)
        elif exp_date == tomorrow:
            expiring_tomorrow.append(batch)

    if not (expired or expiring_today or expiring_tomorrow):
        return

    msg = "🔔 Проверка сроков годности:\n"
    if expired:
        msg += "\n❌ ПРОСРОЧЕНО:\n"
        for b in expired:
            msg += f"• {b['ingredient']} – {b['current_quantity']} {b['unit']} (партия от {b['purchase_date']})\n"
    if expiring_today:
        msg += "\n🔴 Истекает СЕГОДНЯ:\n"
        for b in expiring_today:
            msg += f"• {b['ingredient']} – {b['current_quantity']} {b['unit']}\n"
    if expiring_tomorrow:
        msg += "\n🟠 Истекает ЗАВТРА:\n"
        for b in expiring_tomorrow:
            msg += f"• {b['ingredient']} – {b['current_quantity']} {b['unit']}\n"

    await context.bot.send_message(chat_id=admin_chat_id, text=msg)
async def expiring(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показать продукты с истекающим сроком годности: /expiring [дней]"""
    days = 3
    if context.args:
        try:
            days = int(context.args[0])
        except:
            pass

    today = datetime.now().date()
    deadline = today + timedelta(days=days)
    expiring_list = []
    for batch in batches:
        if not batch.get('is_active', True):
            continue
        if batch.get('opened_date'):
            exp_date = datetime.fromisoformat(batch['expiry_after_open']).date()
        else:
            exp_date = datetime.fromisoformat(batch['expiry_date']).date()
        if exp_date <= deadline:
            expiring_list.append((batch, exp_date))

    if not expiring_list:
        await update.message.reply_text(f"✅ Нет продуктов, у которых срок истекает в ближайшие {days} дней.")
        return

    msg = f"⚠️ Продукты, у которых срок истекает до {deadline}:\n\n"
    for batch, exp_date in sorted(expiring_list, key=lambda x: x[1]):
        days_left = (exp_date - today).days
        if days_left < 0:
            status = "❌ ПРОСРОЧЕНО"
        elif days_left == 0:
            status = "🔴 Истекает сегодня"
        elif days_left == 1:
            status = "🟠 Истекает завтра"
        else:
            status = f"🟡 Осталось {days_left} дн."
        msg += f"• {batch['ingredient']} – {batch['current_quantity']} {batch['unit']} (партия от {batch['purchase_date']}), {status}\n"

    await update.message.reply_text(msg)
async def export_full_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Экспортирует все данные в один Excel-файл с несколькими листами."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
        from io import BytesIO

        wb = openpyxl.Workbook()

        # ========== Лист с ингредиентами ==========
        ws_ing = wb.active
        ws_ing.title = "Ингредиенты"
        headers_ing = ["Название", "Цена (руб)", "Единица", "Остаток"]
        ws_ing.append(headers_ing)
        for col in range(1, len(headers_ing)+1):
            ws_ing.cell(row=1, column=col).font = Font(bold=True)
        for name, data in ingredients.items():
            ws_ing.append([name, data['price'], data['unit'], data.get('stock', 0.0)])

        # ========== Лист с рецептами ==========
        ws_rec = wb.create_sheet("Рецепты")
        headers_rec = ["Название", "Категория", "Тип", "База", "Ингредиенты", "Упаковка", "Работа (ч)", "Наценка %", "Описание"]
        ws_rec.append(headers_rec)
        for col in range(1, len(headers_rec)+1):
            ws_rec.cell(row=1, column=col).font = Font(bold=True)
        for name, data in recipes.items():
            if isinstance(data, dict):
                category = data.get('category', '')
                recipe_type = "весовой" if data.get("type") == "weight" else "штучный"
                base = data.get("base_qty", "")
                ingredients_str = ", ".join([f"{k}: {v}" for k, v in data.get("ingredients", {}).items()])
                packaging = data.get("packaging", "")
                work_hours = data.get("work_hours", "")
                markup = data.get("markup", "")
                description = data.get("description", "")
            else:
                # Старый формат
                category = ''
                recipe_type = "старый"
                base = ""
                ingredients_str = ", ".join([f"{k}: {v}" for k, v in data.items()])
                packaging = work_hours = markup = description = ""
            ws_rec.append([name, category, recipe_type, base, ingredients_str, packaging, work_hours, markup, description])

        # ========== Лист с продажами ==========
        ws_sales = wb.create_sheet("Продажи")
        headers_sales = ["Дата", "Рецепт", "Количество", "Себест. ингр.", "Полная себест.", "Цена продажи", "Прибыль"]
        ws_sales.append(headers_sales)
        for col in range(1, len(headers_sales)+1):
            ws_sales.cell(row=1, column=col).font = Font(bold=True)
        for sale in sales:
            ws_sales.append([
                sale.get('date', ''),
                sale.get('recipe', ''),
                sale.get('quantity', 0),
                sale.get('cost', 0.0),
                sale.get('cost_with_extras', 0.0),
                sale.get('price', ''),
                sale.get('profit', '')
            ])

        # ========== Лист с планами ==========
        ws_plans = wb.create_sheet("Планы")
        headers_plans = ["Дата", "Рецепт", "Количество"]
        ws_plans.append(headers_plans)
        for col in range(1, len(headers_plans)+1):
            ws_plans.cell(row=1, column=col).font = Font(bold=True)
        for plan in plans:
            ws_plans.append([plan.get('date', ''), plan.get('recipe', ''), plan.get('quantity', 0)])

        # ========== Лист с клиентами ==========
        ws_cust = wb.create_sheet("Клиенты")
        headers_cust = ["Имя", "Телефон", "Адрес", "Заметки"]
        ws_cust.append(headers_cust)
        for col in range(1, len(headers_cust)+1):
            ws_cust.cell(row=1, column=col).font = Font(bold=True)
        for name, data in customers.items():
            ws_cust.append([name, data.get('phone', ''), data.get('address', ''), data.get('notes', '')])

        # ========== Лист с заказами ==========
        ws_orders = wb.create_sheet("Заказы")
        headers_orders = ["Клиент", "Рецепт", "Количество", "Дата выполнения", "Статус"]
        ws_orders.append(headers_orders)
        for col in range(1, len(headers_orders)+1):
            ws_orders.cell(row=1, column=col).font = Font(bold=True)
        for order in orders:
            ws_orders.append([
                order.get('customer', ''),
                order.get('recipe', ''),
                order.get('quantity', 0),
                order.get('due_date', ''),
                order.get('status', '')
            ])

        # ========== Лист с партиями (если есть) ==========
        if 'batches' in globals() and batches:
            ws_batches = wb.create_sheet("Партии")
            headers_batches = ["ID", "Ингредиент", "Количество", "Ед.", "Дата закупки", "Срок годности", "Себестоимость", "Поставщик"]
            ws_batches.append(headers_batches)
            for col in range(1, len(headers_batches)+1):
                ws_batches.cell(row=1, column=col).font = Font(bold=True)
            for batch in batches:
                ws_batches.append([
                    batch.get('id', ''),
                    batch.get('ingredient', ''),
                    batch.get('current_quantity', 0),
                    batch.get('unit', ''),
                    batch.get('purchase_date', ''),
                    batch.get('expiry_date', ''),
                    batch.get('cost', 0),
                    batch.get('supplier', '')
                ])

        # Автоподбор ширины колонок для всех листов
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        await update.message.reply_document(
            document=output,
            filename="full_export.xlsx",
            caption="📊 Полный экспорт всех данных (Excel)"
        )
        output.close()

    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка при экспорте: {e}")
async def set_shelf_life(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Установить срок годности для ингредиента: /set_shelf_life ингредиент общий_срок [срок_после_вскрытия]"""
    args = context.args
    if len(args) < 2:
        await update.message.reply_text(
            "Формат: /set_shelf_life ингредиент общий_срок [срок_после_вскрытия]\n"
            "Пример: /set_shelf_life сливки 14 5"
        )
        return

    # Определяем, есть ли второй срок (третий аргумент)
    try:
        # Проверим, является ли последний аргумент числом
        int(args[-1])
        # Если да, то возможно, это общий срок или второй срок.
        if len(args) >= 3:
            try:
                int(args[-2])
                # Если и предпоследний число, значит два числа
                total = int(args[-2])
                after = int(args[-1])
                name_parts = args[:-2]
            except ValueError:
                # Предпоследний не число, значит только общий срок
                total = int(args[-1])
                after = None
                name_parts = args[:-1]
        else:
            # Всего два аргумента: название и общий срок
            total = int(args[-1])
            after = None
            name_parts = args[:-1]
    except ValueError:
        await update.message.reply_text("Сроки должны быть целыми числами (дни).")
        return

    name = ' '.join(name_parts).lower()
    if name not in ingredients:
        await update.message.reply_text(f"Ингредиент '{name}' не найден.")
        return

    # Сохраняем
    ingredients[name]['shelf_life'] = total
    if after is not None:
        ingredients[name]['shelf_life_after_open'] = after
    else:
        ingredients[name]['shelf_life_after_open'] = total
    save_data(ingredients, INGREDIENTS_FILE)

    msg = f"✅ Для '{name}' установлен срок годности: {total} дн."
    if after:
        msg += f" После вскрытия: {after} дн."
    else:
        msg += f" После вскрытия (не задан, используется общий): {total} дн."
    await update.message.reply_text(msg)
async def open_batch(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отметить партию как открытую: /open ингредиент [batch_id]"""
    if len(context.args) < 1:
        await update.message.reply_text(
            "Формат: /open ингредиент [batch_id]\n"
            "Если batch_id не указан, открывается самая старая активная партия этого ингредиента."
        )
        return
    # Разбираем название (может быть из нескольких слов)
    # Последний аргумент может быть batch_id (если передан)
    if len(context.args) == 1:
        name = context.args[0].lower()
        batch_id = None
    else:
        *name_parts, batch_id = context.args
        name = ' '.join(name_parts).lower()

    if name not in ingredients:
        await update.message.reply_text(f"Ингредиент '{name}' не найден.")
        return

    # Находим партии для этого ингредиента, которые ещё не открыты и активны
    candidates = [b for b in batches if b.get('is_active', True) and b['ingredient'] == name and b.get('opened_date') is None]
    if not candidates:
        await update.message.reply_text(f"Нет неоткрытых активных партий для '{name}'.")
        return

    if batch_id:
        # Ищем по id
        batch = next((b for b in candidates if b['id'] == batch_id), None)
        if not batch:
            await update.message.reply_text(f"Партия с id {batch_id} не найдена или уже открыта.")
            return
    else:
        # Выбираем самую старую по сроку годности
        batch = min(candidates, key=lambda b: b['expiry_date'])

    # Устанавливаем дату вскрытия
    opened_date = datetime.now().date()
    batch['opened_date'] = opened_date.isoformat()
    # Вычисляем новый срок после вскрытия
    after_open_days = ingredients[name].get('shelf_life_after_open', ingredients[name].get('shelf_life', 30))
    expiry_after_open = opened_date + timedelta(days=after_open_days)
    batch['expiry_after_open'] = expiry_after_open.isoformat()
    save_batches()

    batch_id = batch.get('id', 'неизвестно')
    await update.message.reply_text(
        f"✅ Партия '{batch_id}' от {batch['purchase_date']} отмечена как открытая.\n"
        f"Новый срок годности: {expiry_after_open} (через {after_open_days} дн. после вскрытия)."
        )
# ========== Команда /price (умная цена) ==========
# ========== Команда /price (умная цена) ==========
# ========== Команда /price (умная цена) ==========
# ========== Команда /price (умная цена) ==========
async def price_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает рекомендованную цену на основе себестоимости и наценок"""
    if not context.args:
        await update.message.reply_text(
            "Введите название рецепта (и количество через пробел, если нужно):\n"
            "Например: меренговый рулет 2"
        )
        return

    # --- Парсим название и количество ---
    args = context.args
    qty = 1.0
    name_parts = []

    for arg in args:
        try:
            qty = float(arg.replace(',', '.'))
        except ValueError:
            name_parts.append(arg)

    name = ' '.join(name_parts).lower()

    if name not in recipes:
        await update.message.reply_text(f"Рецепт '{name}' не найден. Проверьте название.")
        return

    data = recipes[name]

    # --- Расчёт себестоимости ---
    if isinstance(data, dict) and "type" in data:
        ing_dict = data.get("ingredients", {})
        base_qty = data.get("base_qty", 1)
        scale = qty / base_qty if data["type"] == "weight" else qty
    else:
        ing_dict = data.get("ingredients", data) if isinstance(data, dict) else data
        portions = data.get("portions", 1) if isinstance(data, dict) else 1
        scale = qty / portions if portions != 1 else qty

    scaled = {ing: ing_qty * scale for ing, ing_qty in ing_dict.items()}

    total_ing = 0.0
    missing = []
    for ing_name, ing_qty in scaled.items():
        if ing_name in ingredients:
            total_ing += ingredients[ing_name]["price"] * ing_qty
        else:
            missing.append(ing_name)

    if missing:
        await update.message.reply_text(f"❌ Не хватает ингредиентов в базе: {', '.join(missing)}")
        return

    packaging = data.get('packaging', 0.0) if isinstance(data, dict) else 0.0
    work_hours = data.get('work_hours', 0.0) if isinstance(data, dict) else 0.0
    hourly_rate = settings.get('hourly_rate', 0.0)
    work_cost = work_hours * hourly_rate * (qty if 'type' in data and data['type'] == 'pcs' else 1)

    total_cost = total_ing + packaging * (qty if 'type' in data and data['type'] == 'pcs' else 1) + work_cost
    cost_per_unit = total_cost / qty

    is_weight = isinstance(data, dict) and data.get("type") == "weight"
    unit = "кг" if is_weight else "шт"
    unit_display = f" ({cost_per_unit:.0f} руб/{unit})" if qty > 1 else ""

    def round_price(p):
        return int(round(p))

    # --- Наценки ---
    price_min = round_price(total_cost * 1.3)
    price_opt = round_price(total_cost * 1.5)
    price_premium = round_price(total_cost * 1.7)
    price_market = round_price(total_cost * 2.0)

    profit_min = price_min - total_cost
    profit_opt = price_opt - total_cost
    profit_premium = price_premium - total_cost
    profit_market = price_market - total_cost

    margin_min = (profit_min / total_cost) * 100
    margin_opt = (profit_opt / total_cost) * 100
    margin_premium = (profit_premium / total_cost) * 100
    margin_market = (profit_market / total_cost) * 100

    qty_str = str(int(qty)) if qty.is_integer() else f"{qty:.1f}"
    msg = (
        f"🍰 *{name}*\n\n"
        f"Количество: {qty_str} {unit}{unit_display}\n\n"
        f"*Себестоимость:*\n"
        f"{total_cost:.0f} руб ({cost_per_unit:.0f} руб/{unit})\n\n"
        f"💰 *Рекомендованные цены*\n\n"
        f"*Минимальная*\n"
        f"{price_min} руб ({price_min/qty:.0f} / {unit})\n"
        f"Прибыль: {profit_min:.0f} руб · Маржа: {margin_min:.0f}%\n\n"
        f"*Оптимальная* ⭐\n"
        f"{price_opt} руб ({price_opt/qty:.0f} / {unit})\n"
        f"Прибыль: {profit_opt:.0f} руб · Маржа: {margin_opt:.0f}%\n\n"
        f"*Премиум*\n"
        f"{price_premium} руб ({price_premium/qty:.0f} / {unit})\n"
        f"Прибыль: {profit_premium:.0f} руб · Маржа: {margin_premium:.0f}%\n\n"
        f"*Средняя по рынку* (×2)\n"
        f"{price_market} руб ({price_market/qty:.0f} / {unit})\n"
        f"Прибыль: {profit_market:.0f} руб · Маржа: {margin_market:.0f}%"
    )

    # --- Сохраняем для кнопок ---
    context.user_data['price_data'] = {
        'name': name,
        'qty': qty,
        'unit': unit,
        'price_min': price_min,
        'price_opt': price_opt,
        'price_premium': price_premium,
        'price_market': price_market,
        'total_cost': total_cost
    }

    # --- Кнопки ---
    keyboard = [
        [
            InlineKeyboardButton(f"Продать за {price_min}₽", callback_data="price_sell_0"),
            InlineKeyboardButton(f"Продать за {price_opt}₽", callback_data="price_sell_1")
        ],
        [
            InlineKeyboardButton(f"Продать за {price_premium}₽", callback_data="price_sell_2"),
            InlineKeyboardButton(f"Продать за {price_market}₽", callback_data="price_sell_3")
        ],
        [InlineKeyboardButton("❌ Отмена", callback_data="price_cancel")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(msg, parse_mode='Markdown', reply_markup=reply_markup)


async def price_button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "price_cancel":
        await query.edit_message_text("❌ Выбор цены отменён.")
        return

    if data.startswith("price_sell_"):
        try:
            idx = int(data.split('_')[2])
        except Exception:
            await query.edit_message_text("❌ Ошибка выбора.")
            return

        price_data = context.user_data.get('price_data')
        if not price_data:
            await query.edit_message_text("❌ Данные устарели. Повторите /price.")
            return

        name = price_data['name']
        qty = price_data['qty']
        unit = price_data['unit']

        if idx == 0:
            price = price_data['price_min']
            level = "Минимальная"
        elif idx == 1:
            price = price_data['price_opt']
            level = "Оптимальная"
        elif idx == 2:
            price = price_data['price_premium']
            level = "Премиум"
        else:
            price = price_data['price_market']
            level = "Средняя по рынку"

        # Выполняем продажу
        result_msg = await execute_sale(update, context, name, qty, price)

        # Отправляем результат, заменяя сообщение с кнопками
        await query.edit_message_text(result_msg)
    else:
        await query.edit_message_text("❓ Неизвестная команда.")
async def shopping(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показать список закупок на основе заказов и низких остатков. Использование: /shopping [ГГГГ-ММ-ДД]"""
    target_date = None
    days_ahead = 3

    # Пороги низкого остатка для разных единиц
    low_stock_thresholds = {
        'шт': 3,      # для штучных – меньше 3
        'кг': 0.5,    # для кг – меньше 0.5
        'г': 500,     # для граммов – меньше 500 г (0.5 кг)
        'л': 0.5,     # для литров – меньше 0.5 л
        'мл': 500     # для миллилитров – меньше 500 мл (0.5 л)
    }

    if context.args:
        try:
            target_date = datetime.strptime(context.args[0], "%Y-%m-%d").date()
        except ValueError:
            await update.message.reply_text("Неверный формат даты. Используйте ГГГГ-ММ-ДД")
            return
    else:
        target_date = datetime.now().date() + timedelta(days=days_ahead)

    today = datetime.now().date()
    relevant_orders = []
    for order in orders:
        order_date = datetime.fromisoformat(order['due_date']).date()
        if today <= order_date <= target_date:
            relevant_orders.append(order)

    needs = {}
    if relevant_orders:
        for order in relevant_orders:
            recipe_name = order['recipe']
            qty = order['quantity']
            if recipe_name not in recipes:
                continue
            data = recipes[recipe_name]
            if isinstance(data, dict) and "ingredients" in data:
                ing_dict = data["ingredients"]
            else:
                ing_dict = data
            for ing_name, ing_qty in ing_dict.items():
                need = ing_qty * qty
                needs[ing_name] = needs.get(ing_name, 0.0) + need

        deficit = {}
        for ing_name, need in needs.items():
            stock = ingredients.get(ing_name, {}).get('stock', 0.0)
            need_deficit = need - stock
            if need_deficit > 0:
                deficit[ing_name] = need_deficit
    else:
        deficit = {}

    # Определяем ингредиенты с низким остатком с учётом единиц
    low_stock = {}
    for ing_name, data in ingredients.items():
        stock = data.get('stock', 0.0)
        unit = data.get('unit', 'кг')
        threshold = low_stock_thresholds.get(unit, 0.5)  # по умолчанию 0.5

        if stock < threshold:
            low_stock[ing_name] = (stock, unit, threshold)

    # Убираем из low_stock те, что уже есть в deficit
    for ing_name in deficit.keys():
        low_stock.pop(ing_name, None)

    if not deficit and not low_stock:
        await update.message.reply_text("✅ У вас достаточно ингредиентов для заказов, и все остатки в норме.")
        return

    msg = "🛒 Список покупок\n\n"

    if deficit:
        msg += "Для заказов до {}:\n".format(target_date)
        for ing_name, need in sorted(deficit.items()):
            unit = ingredients.get(ing_name, {}).get('unit', '')
            msg += f"• {ing_name}: {need:.2f} {unit}\n"
        msg += "\n"

    if low_stock:
        msg += f"⚠️ Низкий остаток:\n"
        for ing_name, (stock, unit, threshold) in sorted(low_stock.items()):
            msg += f"• {ing_name}: {stock:.2f} {unit} (менее {threshold} {unit})\n"

    await update.message.reply_text(msg)
async def help_button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    # Словарь соответствия callback_data и текстов разделов
    help_sections = {
        "help_main": HELP_MAIN,
        "help_ingredients": HELP_INGREDIENTS,
        "help_recipes": HELP_RECIPES,
        "help_sales": HELP_SALES,
        "help_writeoffs": HELP_WRITEOFFS,
        "help_customers": HELP_CUSTOMERS,
        "help_plans": HELP_PLANS,
        "help_import_export": HELP_IMPORT_EXPORT,
        "help_advanced": HELP_ADVANCED,
    }

    if data in help_sections:
        # Отправляем текст раздела и кнопку «Назад»
        keyboard = [[InlineKeyboardButton("« Назад к разделам", callback_data="help_back")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(help_sections[data], reply_markup=reply_markup)
    elif data == "help_back":
        # Возвращаем главное меню помощи
        await help_command(update, context)
    else:
        # Неизвестный callback – игнорируем или логируем
        print(f"Неизвестный help callback: {data}")
def get_main_keyboard():
    """Главное меню с разделами"""
    keyboard = [
        [KeyboardButton("📦 Ингредиенты")],
        [KeyboardButton("🍰 Рецепты")],
        [KeyboardButton("💰 Продажи")],
        [KeyboardButton("📊 Аналитика")],
        [KeyboardButton("🛒 Закупки")],
        [KeyboardButton("👥 Клиенты")],
        [KeyboardButton("❓ Помощь")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
def get_ingredients_submenu():
    """Подменю раздела Ингредиенты"""
    keyboard = [
        [KeyboardButton("➕ Добавить ингредиент")],
        [KeyboardButton("📋 Список ингредиентов")],
        [KeyboardButton("🔄 Обновить цену")],
        [KeyboardButton("📦 Закупка (партия)")],
        [KeyboardButton("⏰ Сроки годности")],
        [KeyboardButton("« Назад")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_recipes_submenu():
    """Подменю раздела Рецепты"""
    keyboard = [
        [KeyboardButton("➕ Новый рецепт")],
        [KeyboardButton("📋 Мои рецепты")],
        [KeyboardButton("🔍 Показать рецепт")],
        [KeyboardButton("⚖️ Пересчитать")],
        [KeyboardButton("💰 Себестоимость")],
        [KeyboardButton("📈 Прайс-лист")],
        [KeyboardButton("« Назад")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_sales_submenu():
    keyboard = [
        [KeyboardButton("📦 Новый заказ")],          # вместо "💰 Продать"
        [KeyboardButton("💵 Рекомендованная цена")],
        [KeyboardButton("📊 Прибыль")],              # вместо "📊 Статистика"
        [KeyboardButton("🏆 Популярные")],
        [KeyboardButton("🔄 Возврат")],
        [KeyboardButton("❌ Списание")],
        [KeyboardButton("« Назад")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
def get_analytics_submenu():
    """Подменю раздела Аналитика"""
    keyboard = [
        [KeyboardButton("📈 Прибыль за месяц")],
        [KeyboardButton("📉 Самые прибыльные")],
        [KeyboardButton("📊 Отчёт Excel")],
        [KeyboardButton("« Назад")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_purchases_submenu():
    """Подменю раздела Закупки"""
    keyboard = [
        [KeyboardButton("🛒 Список покупок")],
        [KeyboardButton("📦 Запланировать")],
        [KeyboardButton("⏳ Истекающие сроки")],
        [KeyboardButton("« Назад")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_customers_submenu():
    """Подменю раздела Клиенты"""
    keyboard = [
        [KeyboardButton("➕ Новый клиент")],
        [KeyboardButton("📅 Создать заказ")],
        [KeyboardButton("📋 Заказы на дату")],
        [KeyboardButton("🔔 Напоминания")],
        [KeyboardButton("« Назад")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
# ========== Онбординг (обучение нового пользователя) ==========
# ========== Онбординг (обучение нового пользователя) ==========
async def onboarding_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начинает онбординг"""
    if context.user_data.get('onboarding_complete'):
        await show_main_menu(update, context, first_time=False)
        return ConversationHandler.END

    context.user_data["onboarding"] = True

    await update.message.reply_text(
        "👋 Привет! Я помогу тебе вести учёт твоего кондитерского производства.\n\n"
        "Давай быстро настроим систему. Это займёт пару минут.\n"
        "Любой шаг можно пропустить."
    )

    return await onboarding_step1(update, context)


async def onboarding_step1(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [KeyboardButton("➕ Добавить ингредиент")],
        [KeyboardButton("⏩ Пропустить")]
    ]

    await update.message.reply_text(
        "📦 **Шаг 1 из 4**\n\n"
        "Добавим ингредиенты.\n\n"
        "Пример:\n"
        "`мука 50 кг`\n"
        "`масло 209.99 180 г`",
        parse_mode='Markdown',
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )

    return ONBOARDING_INGREDIENT


async def onboarding_ingredient(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()

    if text == "⏩ Пропустить":
        return await onboarding_step2(update, context)

    if text == "➕ Добавить ингредиент" or text == "➕ Добавить ещё":
        await update.message.reply_text(
            "Отправь ингредиент:\n\n"
            "`мука 50 кг`\n"
            "`масло 209.99 180 г`",
            parse_mode="Markdown"
        )
        return ONBOARDING_INGREDIENT

    if text == "⏩ Дальше":
        return await onboarding_step2(update, context)

    args = text.split()

    if len(args) >= 3:
        context.args = args
        await add_ingredient(update, context)

        keyboard = [
            [KeyboardButton("➕ Добавить ещё"), KeyboardButton("⏩ Дальше")]
        ]

        await update.message.reply_text(
            "✅ Ингредиент добавлен!",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )

        return ONBOARDING_INGREDIENT

    await update.message.reply_text(
        "❗ Неверный формат.\n\n"
        "Пример:\n"
        "`мука 50 кг`\n"
        "`масло 209.99 180 г`",
        parse_mode="Markdown"
    )

    return ONBOARDING_INGREDIENT


async def onboarding_step2(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [KeyboardButton("🍰 Создать рецепт")],
        [KeyboardButton("⏩ Пропустить")]
    ]

    await update.message.reply_text(
        "🍰 **Шаг 2 из 4**\n\n"
        "Создадим рецепт.\n\n"
        "Пример:\n"
        "`/add_recipe2 омлет штук 1: яйца 2, молоко 0.1`",
        parse_mode='Markdown',
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )

    return ONBOARDING_RECIPE


async def onboarding_recipe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()

    if text == "⏩ Пропустить":
        return await onboarding_step3(update, context)

    if text == "⏩ Дальше":
        return await onboarding_step3(update, context)

    if text.startswith('/add_recipe2'):
        context.args = text.split()[1:]
        await add_recipe_scaled(update, context)

        keyboard = [
            [KeyboardButton("⏩ Дальше")]
        ]

        await update.message.reply_text(
            "✅ Рецепт создан!",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )

        return ONBOARDING_RECIPE

    await update.message.reply_text(
        "Отправь команду:\n"
        "`/add_recipe2 название тип количество: ингредиенты`",
        parse_mode="Markdown"
    )

    return ONBOARDING_RECIPE


async def onboarding_step3(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [KeyboardButton("💰 Рассчитать цену")],
        [KeyboardButton("⏩ Пропустить")]
    ]

    await update.message.reply_text(
        "💰 **Шаг 3 из 4**\n\n"
        "Попробуем рассчитать цену.\n\n"
        "Пример:\n"
        "`/price омлет 2`",
        parse_mode='Markdown',
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )

    return ONBOARDING_PRICE


async def onboarding_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()

    if text == "⏩ Пропустить":
        return await onboarding_step4(update, context)

    if text == "⏩ Дальше":
        return await onboarding_step4(update, context)

    if text.startswith('/price'):
        context.args = text.split()[1:]
        await price_command(update, context)

        keyboard = [
            [KeyboardButton("⏩ Дальше")]
        ]

        await update.message.reply_text(
            "✅ Видишь — бот посчитал себестоимость и цену!",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )

        return ONBOARDING_PRICE

    await update.message.reply_text(
        "Отправь команду:\n"
        "`/price название 2`",
        parse_mode="Markdown"
    )

    return ONBOARDING_PRICE


async def onboarding_step4(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["onboarding"] = False
    context.user_data["onboarding_complete"] = True

    await update.message.reply_text(
        "🎉 **Готово!**\n\n"
        "Теперь ты можешь использовать бота.\n\n"
        "Главные команды:\n"
        "/ingredients\n"
        "/recipes\n"
        "/price\n"
        "/profit",
        parse_mode='Markdown',
        reply_markup=get_main_keyboard()
    )

    return ConversationHandler.END


async def onboarding_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["onboarding"] = False

    await update.message.reply_text(
        "❌ Обучение отменено.\n"
        "Запустить снова можно командой /onboarding",
        reply_markup=get_main_keyboard()
    )

    return ConversationHandler.END
    await update.message.reply_text("Я не понял команду. Попробуйте /help")
def main():
    import os
    from datetime import time
    from telegram.ext import (
        Application,
        CommandHandler,
        MessageHandler,
        CallbackQueryHandler,
        ConversationHandler,
        filters,
    )
    from telegram.request import HTTPXRequest

    print("🔥 Запуск main()")

    TOKEN = os.environ.get("BOT_TOKEN")
    if not TOKEN:
        print("❌ BOT_TOKEN не найден")
        return

    print("✅ Токен загружен")

    request = HTTPXRequest(
        connection_pool_size=20,
        connect_timeout=60,
        read_timeout=60,
        write_timeout=60,
        pool_timeout=60,
    )

    application = Application.builder().token(TOKEN).request(request).build()

    # --- Загрузка данных ---
    global ingredients, recipes, settings

    ingredients.clear()
    ingredients.update(load_data(INGREDIENTS_FILE))

    recipes.clear()
    recipes.update(load_data(RECIPES_FILE))

    load_settings()
    load_sales()
    load_plans()
    load_customers()
    load_writeoffs()
    load_orders()
    load_price_history()
    load_batches()

    print("✅ Данные загружены")

    # -------------------------------------------------
    # Диалог импорта рецепта
    # -------------------------------------------------

    import_conv = ConversationHandler(
        entry_points=[CommandHandler("import_recipe", import_recipe_start)],
        states={
            WAITING_RECIPE_TEXT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_recipe_text)
            ],
            WAITING_INGREDIENT_PRICE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_ingredient_price)
            ],
            WAITING_RECIPE_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_recipe_name)
            ],
            WAITING_RECIPE_TYPE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_recipe_type)
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    application.add_handler(import_conv)

    # -------------------------------------------------
    # Онбординг
    # -------------------------------------------------

    onboarding_handler = ConversationHandler(
        entry_points=[
            CommandHandler("start", onboarding_start),
            CommandHandler("onboarding", onboarding_start),
        ],
        states={
            ONBOARDING_START: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, onboarding_step1)
            ],
            ONBOARDING_INGREDIENT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, onboarding_ingredient)
            ],
            ONBOARDING_RECIPE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, onboarding_recipe)
            ],
            ONBOARDING_PRICE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, onboarding_price)
            ],
            ONBOARDING_FINISH: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, onboarding_step4)
            ],
        },
        fallbacks=[CommandHandler("cancel", onboarding_cancel)],
    )

    application.add_handler(onboarding_handler)

    # -------------------------------------------------
    # Команды
    # -------------------------------------------------

    commands = [
    ("menu", menu),
    ("add_ingredient", add_ingredient),
    ("ingredients", show_ingredients),
    ("add_recipe", add_recipe),
    ("add_recipe2", add_recipe_scaled),
    ("recipes", list_recipes),
    ("calculate", calculate_cost),
    ("scale", scale_recipe),
    ("remove_ingredient", remove_ingredient),
    ("remove_recipe", remove_recipe),
    ("update_price", update_price),
    ("export", export_data),
    ("help", help_command),
    ("delete_recipes", delete_all_recipes),
    ("set_description", set_description),
    ("show_recipe", show_recipe),
    ("set_hourly_rate", set_hourly_rate),
    ("set_packaging", set_packaging),
    ("set_work_hours", set_work_hours),
    ("set_markup", set_markup),
    ("price_list", price_list),
    ("parse", parse_recipe),
    ("set_category", set_category),
    ("categories", list_categories),
    ("stats", stats),
    ("popular", popular),
    ("add_stock", add_stock),
    ("stock", show_stock),
    ("low_stock", low_stock),
    ("plan", plan_recipe),
    ("export_full", export_full),
    ("use", use_recipe),
    ("export_xlsx", export_xlsx),
    ("report_xlsx", report_xlsx),
    ("add_customer", add_customer),
    ("preorder", create_order),          # старая команда предзаказа переименована
    ("orders", list_orders),
    ("remind", remind_orders),
    ("set_admin", set_admin),
    ("write_off", write_off),
    ("refund", refund),
    ("price_history", price_history_cmd),
    ("purchase", purchase),
    ("expiring", expiring),
    ("export_full_excel", export_full_excel),
    ("set_shelf_life", set_shelf_life),
    ("open", open_batch),
    ("price", price_command),
    ("shopping", shopping),               # команда для списка покупок
    ("order", order_command),              # новая команда учёта заказов
    ("profit", profit_command),             # команда прибыли
]
    for cmd, func in commands:
        application.add_handler(CommandHandler(cmd, func))
    # -------------------------------------------------
    # Callback кнопки
    # -------------------------------------------------

    application.add_handler(
        CallbackQueryHandler(price_button_handler, pattern="^price_")
    )

    application.add_handler(
        CallbackQueryHandler(help_button_handler, pattern="^help_")
    )

    application.add_handler(
        CallbackQueryHandler(button_handler)
    )

    # -------------------------------------------------
    # Планировщик
    # -------------------------------------------------

    job_queue = application.job_queue

    if job_queue:
        job_queue.run_daily(
            scheduled_remind,
            time=time(hour=19, minute=0),
        )

        job_queue.run_daily(
            expiry_check,
            time=time(hour=8, minute=0),
        )

    # -------------------------------------------------
    # Текстовые сообщения
    # -------------------------------------------------

    application.add_handler(
        MessageHandler(filters.TEXT & ~filters.COMMAND, handle_menu_buttons),
        group=0,
    )

    print("🚀 Бот запущен")

    application.run_polling()


if __name__ == "__main__":
    main()
